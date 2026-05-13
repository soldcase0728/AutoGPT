"""Convert each sheet of student_offer.xlsx to a CSV for downstream analysis."""
import csv
import openpyxl

wb = openpyxl.load_workbook("student_offer.xlsx", data_only=True)

OUT = {
    "Student Offer Schedule": "student_offer.csv",
    "Waitlist + Backups": "waitlist.csv",
    "Capacity Report": "capacity_report.csv",
    "Summary Dashboard": "summary_dashboard.csv",
}

for sheet, path in OUT.items():
    ws = wb[sheet]
    with open(path, "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        for row in ws.iter_rows(values_only=True):
            w.writerow(["" if v is None else v for v in row])
    print(f"wrote {path} ({ws.max_row} rows)")
