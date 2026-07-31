"""Excel / CSV / Markdown report generation.

Every workbook produced here carries three standard tabs in addition to
its data:

* **Data Dictionary** - what each column means and, critically, whether
  the value is OBSERVED (read directly off the file or page), CALCULATED
  (arithmetic over observed values), or INFERRED (a heuristic judgement).
* **Methodology and Limitations** - how the numbers were produced and what
  they cannot support.
* **Run Statistics** - the run that produced the workbook.

Confidentiality
---------------
No report ever contains extracted document text. Issue-tag rows carry a
structural locator (page and offset), not an excerpt, unless a reviewer
has deliberately raised ``issue_tags.context_chars`` in config.yaml.
"""

from __future__ import annotations

import csv
import json
import sqlite3
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any, Iterable, Mapping, Sequence

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from src.config import Config
from src.database import Database
from src.logging_setup import get_logger
from src.version import APP_NAME, APP_VERSION

LOG = get_logger("reports")

HEADER_FILL = PatternFill("solid", fgColor="DDDDDD")
HEADER_FONT = Font(bold=True)

#: Provenance vocabulary. Used verbatim in every Data Dictionary tab.
OBSERVED = "OBSERVED"
CALCULATED = "CALCULATED"
INFERRED = "INFERRED"

DERIVED_NOTICE = (
    "All values in this workbook are DERIVED REVIEW METADATA produced by "
    f"{APP_NAME} v{APP_VERSION} from the PDFs themselves. They are NOT original "
    "or native metadata from the producing party, and no load file, DAT, OPT, "
    "or native-file index was available for this production."
)

LIMITATIONS = [
    DERIVED_NOTICE,
    "Bates numbers parsed from filenames are preliminary. A single PDF in this "
    "production may span several Bates-stamped pages, so a numeric jump between "
    "consecutive filenames often reflects the page count of the preceding "
    "document rather than a missing document.",
    "Document-type classification, priority scores, and document-family links "
    "are heuristic. Each carries an explicit confidence value; treat anything "
    "below 'Confirmed' as a lead to verify, not a finding.",
    "Issue tags are a screening tool only. A tag means a configured term "
    "appeared in extracted text. It is not a legal conclusion, a privilege "
    "determination, or evidence of anything.",
    "A high priority score means 'review this early'. It does not indicate "
    "malpractice, liability, or any other legal conclusion.",
    "Text extraction quality varies. Image-only pages contribute no text until "
    "OCR runs, and OCR output contains recognition errors. Fields derived from "
    "OCR text are correspondingly less reliable.",
    "Custodian is populated only where the production itself shows one. Folder "
    "structure in a synced Drive is not evidence of custody and is not used.",
    "No accuracy claim is perfect. Measured pilot accuracy and known failure "
    "modes are reported in the phase summary documents.",
    "Source files were never renamed, moved, altered, annotated, OCR'd in "
    "place, combined, or deleted. The source folder is treated as read-only.",
]


@dataclass(frozen=True)
class ColumnSpec:
    """One column in a report, with its provenance and description."""

    name: str
    provenance: str
    description: str
    width: int = 18
    wrap: bool = False


def _write_header(sheet: Worksheet, columns: Sequence[str]) -> None:
    """Write and style the header row."""
    for index, name in enumerate(columns, start=1):
        cell = sheet.cell(row=1, column=index, value=name)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(vertical="center", wrap_text=True)


def _autosize(
    sheet: Worksheet,
    columns: Sequence[str],
    rows: Sequence[Mapping[str, Any]],
    *,
    max_width: int = 60,
    wrap_columns: Sequence[str] = (),
) -> None:
    """Set readable column widths and enable wrapping where configured."""
    sample = rows[:500]
    for index, name in enumerate(columns, start=1):
        longest = len(str(name))
        for row in sample:
            value = row.get(name)
            if value is not None:
                longest = max(longest, min(len(str(value)), max_width))
        letter = get_column_letter(index)
        sheet.column_dimensions[letter].width = min(max(10, longest + 2), max_width)
        if name in wrap_columns:
            for cell in sheet[letter]:
                if cell.row > 1:
                    cell.alignment = Alignment(wrap_text=True, vertical="top")


def write_sheet(
    workbook: Workbook,
    title: str,
    columns: Sequence[str],
    rows: Sequence[Mapping[str, Any]],
    *,
    config: Config | None = None,
    hyperlink_column: str | None = None,
    hyperlink_target_key: str = "_abs_path",
) -> Worksheet:
    """Add a formatted data sheet to a workbook.

    Args:
        workbook: Destination workbook.
        title: Sheet name (Excel truncates beyond 31 characters).
        columns: Ordered column names.
        rows: Row mappings keyed by column name. Keys beginning with ``_``
            are treated as metadata and are not written as cells.
        config: Used for width, wrap, freeze, filter, and hyperlink
            settings. Sensible defaults apply when omitted.
        hyperlink_column: Column whose cells become clickable local-file
            links.
        hyperlink_target_key: Row key holding the absolute path to link to.

    Returns:
        The created worksheet.
    """
    settings = config.section("reports") if config else {}
    max_width = int(settings.get("max_column_width", 60))
    wrap_columns = list(settings.get("wrap_text_columns", []))
    use_links = bool(settings.get("hyperlinks", True))

    sheet = workbook.create_sheet(title=title[:31])
    _write_header(sheet, columns)

    for row_index, row in enumerate(rows, start=2):
        for column_index, name in enumerate(columns, start=1):
            value = row.get(name)
            if isinstance(value, (list, tuple, set)):
                value = "; ".join(str(item) for item in value)
            elif isinstance(value, dict):
                value = json.dumps(value, default=str)
            cell = sheet.cell(row=row_index, column=column_index, value=value)
            if (
                use_links
                and hyperlink_column
                and name == hyperlink_column
                and row.get(hyperlink_target_key)
            ):
                target = Path(str(row[hyperlink_target_key])).as_uri()
                cell.hyperlink = target
                cell.style = "Hyperlink"

    if settings.get("freeze_header", True) if settings else True:
        sheet.freeze_panes = "A2"
    if (settings.get("autofilter", True) if settings else True) and rows:
        sheet.auto_filter.ref = (
            f"A1:{get_column_letter(len(columns))}{len(rows) + 1}"
        )
    _autosize(sheet, columns, rows, max_width=max_width, wrap_columns=wrap_columns)
    return sheet


def add_data_dictionary(workbook: Workbook, specs: Sequence[ColumnSpec]) -> None:
    """Add the standard Data Dictionary tab."""
    rows = [
        {
            "Column": spec.name,
            "Data category": spec.provenance,
            "Description": spec.description,
        }
        for spec in specs
    ]
    rows.extend(
        [
            {"Column": "", "Data category": "", "Description": ""},
            {
                "Column": "OBSERVED",
                "Data category": "definition",
                "Description": (
                    "Read directly from the file or from text present on a page - "
                    "file size, page count, a Bates stamp visible on the page."
                ),
            },
            {
                "Column": "CALCULATED",
                "Data category": "definition",
                "Description": (
                    "Arithmetic over observed values - an end Bates derived from a "
                    "begin Bates plus a page count, a hash, a count, a score."
                ),
            },
            {
                "Column": "INFERRED",
                "Data category": "definition",
                "Description": (
                    "A heuristic judgement - document type, family relationships, "
                    "issue tags. Always accompanied by a confidence value. Never "
                    "native metadata."
                ),
            },
        ]
    )
    sheet = workbook.create_sheet("Data Dictionary")
    columns = ["Column", "Data category", "Description"]
    _write_header(sheet, columns)
    for row_index, row in enumerate(rows, start=2):
        for column_index, name in enumerate(columns, start=1):
            sheet.cell(row=row_index, column=column_index, value=row.get(name))
    sheet.column_dimensions["A"].width = 34
    sheet.column_dimensions["B"].width = 16
    sheet.column_dimensions["C"].width = 100
    for cell in sheet["C"]:
        cell.alignment = Alignment(wrap_text=True, vertical="top")
    sheet.freeze_panes = "A2"


def add_methodology(workbook: Workbook, extra_notes: Sequence[str] = ()) -> None:
    """Add the standard Methodology and Limitations tab."""
    sheet = workbook.create_sheet("Methodology and Limitations")
    _write_header(sheet, ["#", "Methodology and limitations"])
    entries = [*LIMITATIONS, *extra_notes]
    for index, note in enumerate(entries, start=1):
        sheet.cell(row=index + 1, column=1, value=index)
        cell = sheet.cell(row=index + 1, column=2, value=note)
        cell.alignment = Alignment(wrap_text=True, vertical="top")
    sheet.column_dimensions["A"].width = 5
    sheet.column_dimensions["B"].width = 120
    sheet.freeze_panes = "A2"


def add_run_statistics(workbook: Workbook, stats: Mapping[str, Any]) -> None:
    """Add the standard Run Statistics tab."""
    sheet = workbook.create_sheet("Run Statistics")
    _write_header(sheet, ["Statistic", "Value"])
    row_index = 2
    for key, value in stats.items():
        if isinstance(value, Mapping):
            for sub_key, sub_value in value.items():
                sheet.cell(row=row_index, column=1, value=f"{key}.{sub_key}")
                sheet.cell(row=row_index, column=2, value=sub_value)
                row_index += 1
            continue
        sheet.cell(row=row_index, column=1, value=key)
        sheet.cell(
            row=row_index,
            column=2,
            value=value if isinstance(value, (str, int, float, type(None))) else str(value),
        )
        row_index += 1
    sheet.column_dimensions["A"].width = 40
    sheet.column_dimensions["B"].width = 60
    sheet.freeze_panes = "A2"


def finalise_workbook(
    workbook: Workbook,
    path: Path,
    *,
    specs: Sequence[ColumnSpec],
    stats: Mapping[str, Any],
    extra_notes: Sequence[str] = (),
) -> Path:
    """Attach the standard tabs, drop the default sheet, and save."""
    add_data_dictionary(workbook, specs)
    add_methodology(workbook, extra_notes)
    add_run_statistics(workbook, stats)
    default = workbook["Sheet"] if "Sheet" in workbook.sheetnames else None
    if default is not None:
        workbook.remove(default)
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)
    LOG.info("Report written", extra={"report": path.name, "path": str(path)})
    return path


def write_csv(path: Path, columns: Sequence[str], rows: Sequence[Mapping[str, Any]]) -> Path:
    """Write a CSV alongside a workbook, excluding metadata keys."""
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=list(columns), extrasaction="ignore")
        writer.writeheader()
        for row in rows:
            writer.writerow(
                {
                    key: ("; ".join(str(item) for item in value)
                          if isinstance(value, (list, tuple, set)) else value)
                    for key, value in row.items()
                    if key in set(columns)
                }
            )
    LOG.info("CSV written", extra={"report": path.name, "rows": len(rows)})
    return path


# ---------------------------------------------------------------------------
# Phase 1 reports
# ---------------------------------------------------------------------------

INVENTORY_COLUMNS: tuple[ColumnSpec, ...] = (
    ColumnSpec("Filename", OBSERVED, "Base filename exactly as it appears on disk.", 34),
    ColumnSpec("Relative path", OBSERVED, "Path relative to the source production folder.", 40),
    ColumnSpec("Google Drive local path", OBSERVED, "Absolute local path of the synced file.", 50),
    ColumnSpec("Open file", CALCULATED, "Clickable file:// link built from the local path.", 14),
    ColumnSpec("File type", OBSERVED, "Lower-cased file extension.", 10),
    ColumnSpec("File size (bytes)", OBSERVED, "On-disk size reported by the filesystem.", 16),
    ColumnSpec("Created", OBSERVED, "Filesystem creation time where the OS records one (macOS birth time). Blank on filesystems that do not track it - never substituted with another timestamp.", 22),
    ColumnSpec("Modified", OBSERVED, "Filesystem modification time.", 22),
    ColumnSpec("SHA-256", CALCULATED, "SHA-256 digest of the complete file bytes.", 26),
    ColumnSpec("Page count", OBSERVED, "Page count read from the PDF structure. Blank when the file could not be opened.", 12),
    ColumnSpec("Encrypted", OBSERVED, "Yes when the PDF is password-protected. No password was ever supplied or guessed.", 12),
    ColumnSpec("Corrupt or unreadable", OBSERVED, "Yes when the PDF failed to parse.", 20),
    ColumnSpec("Cloud placeholder", OBSERVED, "Yes when the file was not materialised locally by Google Drive. Re-run after it downloads.", 18),
    ColumnSpec("Bates from filename", OBSERVED, "Bates value parsed from the filename only. Not read from the page.", 22),
    ColumnSpec("Bates number", CALCULATED, "Numeric portion of the filename Bates value.", 14),
    ColumnSpec("Filename compliant", CALCULATED, "Yes when the filename matches the expected production pattern.", 18),
    ColumnSpec("Duplicate filename group", CALCULATED, "Set when another file shares this exact basename.", 24),
    ColumnSpec("Exact duplicate group", CALCULATED, "Set when another file has an identical SHA-256.", 22),
    ColumnSpec("Processing status", CALCULATED, "Phase 1 outcome for this file.", 18),
    ColumnSpec("Notes", INFERRED, "Structural observations, e.g. a non-standard Bates width.", 40, wrap=True),
    ColumnSpec("Error", OBSERVED, "Error text captured when the file could not be read.", 40, wrap=True),
)


def _yes_no(value: Any) -> str:
    """Render a truthy database flag as an explicit Yes/No string.

    Text rather than colour: the reports must not depend on colour to
    carry meaning.
    """
    return "Yes" if value else "No"


def _inventory_rows(database: Database) -> list[dict[str, Any]]:
    """Build inventory report rows by streaming the files table."""
    rows: list[dict[str, Any]] = []
    for record in database.iter_files():
        rows.append(
            {
                "Filename": record["filename"],
                "Relative path": record["rel_path"],
                "Google Drive local path": record["abs_path"],
                "Open file": "Open",
                "File type": record["extension"],
                "File size (bytes)": record["file_size"],
                "Created": record["created_ts"],
                "Modified": record["modified_ts"],
                "SHA-256": record["sha256"],
                "Page count": record["page_count"],
                "Encrypted": _yes_no(record["is_encrypted"]),
                "Corrupt or unreadable": _yes_no(record["is_corrupt"]),
                "Cloud placeholder": _yes_no(record["is_cloud_placeholder"]),
                "Bates from filename": record["bates_filename_raw"],
                "Bates number": record["bates_filename_num"],
                "Filename compliant": _yes_no(record["filename_compliant"]),
                "Duplicate filename group": record["dup_filename_group"],
                "Exact duplicate group": record["exact_dup_group"],
                "Processing status": record["processing_status"],
                "Notes": record["filename_notes"],
                "Error": record["error_message"] or record["pdf_open_error"],
                "_abs_path": record["abs_path"],
            }
        )
    return rows


def write_phase1_reports(
    config: Config, database: Database, stats: Mapping[str, Any]
) -> dict[str, Path]:
    """Generate every Phase 1 deliverable.

    Args:
        config: Loaded configuration.
        database: Populated processing-state database.
        stats: Aggregate statistics from :func:`inventory.run_inventory`.

    Returns:
        Mapping of report key to written path.
    """
    from src import bates as bates_mod

    reports_dir = config.paths.reports_dir
    reports_dir.mkdir(parents=True, exist_ok=True)
    written: dict[str, Path] = {}
    columns = [spec.name for spec in INVENTORY_COLUMNS]
    rows = _inventory_rows(database)

    # 01 - File inventory
    workbook = Workbook()
    write_sheet(
        workbook,
        "File Inventory",
        columns,
        rows,
        config=config,
        hyperlink_column="Open file",
    )
    written["01_File_Inventory.xlsx"] = finalise_workbook(
        workbook,
        reports_dir / "01_File_Inventory.xlsx",
        specs=INVENTORY_COLUMNS,
        stats=stats,
    )
    written["01_File_Inventory.csv"] = write_csv(
        reports_dir / "01_File_Inventory.csv", columns, rows
    )

    # 02 - Bates structural report
    numbers = [
        int(record["bates_filename_num"])
        for record in database.query(
            "SELECT bates_filename_num FROM files WHERE bates_filename_num IS NOT NULL"
        )
    ]
    gaps = bates_mod.find_filename_gaps(numbers)
    gap_rows = [
        {
            "Gap after Bates": bates_mod.format_bates(gap.after),
            "Gap before Bates": bates_mod.format_bates(gap.before),
            "Missing numbers": gap.label,
            "Count missing": gap.missing_count,
            "Status": "PRELIMINARY - filename sequence only",
            "Basis": gap.basis,
        }
        for gap in gaps
    ]
    malformed_rows = [
        {
            "Filename": record["filename"],
            "Relative path": record["rel_path"],
            "Open file": "Open",
            "Bates from filename": record["bates_filename_raw"],
            "Issue": (
                "No Bates pattern found in the filename"
                if record["bates_filename_num"] is None
                else "Filename does not match the compliant naming pattern"
            ),
            "Notes": record["filename_notes"],
            "_abs_path": record["abs_path"],
        }
        for record in database.iter_files(
            where="is_pdf = 1 AND (bates_filename_num IS NULL OR filename_compliant = 0)"
        )
    ]

    workbook = Workbook()
    write_sheet(
        workbook,
        "Preliminary Bates Gaps",
        [
            "Gap after Bates",
            "Gap before Bates",
            "Missing numbers",
            "Count missing",
            "Status",
            "Basis",
        ],
        gap_rows,
        config=config,
    )
    write_sheet(
        workbook,
        "Malformed Filenames",
        ["Filename", "Relative path", "Open file", "Bates from filename", "Issue", "Notes"],
        malformed_rows,
        config=config,
        hyperlink_column="Open file",
    )
    write_sheet(
        workbook,
        "Bates Range",
        [
            "Statistic",
            "Value",
        ],
        [
            {"Statistic": "Minimum apparent Bates (filename)", "Value": stats.get("min_bates")},
            {"Statistic": "Maximum apparent Bates (filename)", "Value": stats.get("max_bates")},
            {"Statistic": "Files with a parsed Bates number", "Value": stats.get("bates_parsed")},
            {"Statistic": "PDFs with no Bates in the filename", "Value": stats.get("bates_missing")},
            {"Statistic": "Non-compliant filenames", "Value": stats.get("noncompliant_filenames")},
            {"Statistic": "Preliminary gap runs", "Value": stats.get("filename_gap_runs")},
            {
                "Statistic": "Preliminary missing numbers",
                "Value": stats.get("filename_gap_missing_numbers"),
            },
        ],
        config=config,
    )
    written["02_Bates_Structural_Report.xlsx"] = finalise_workbook(
        workbook,
        reports_dir / "02_Bates_Structural_Report.xlsx",
        specs=(
            ColumnSpec("Gap after Bates", OBSERVED, "Last filename Bates value before the gap.", 20),
            ColumnSpec("Gap before Bates", OBSERVED, "First filename Bates value after the gap.", 20),
            ColumnSpec("Missing numbers", CALCULATED, "Range of Bates numbers no filename claims.", 22),
            ColumnSpec("Count missing", CALCULATED, "Size of the gap.", 14),
            ColumnSpec("Status", CALCULATED, "Always PRELIMINARY at Phase 1.", 34),
            ColumnSpec("Basis", CALCULATED, "Why this gap may not be a real gap.", 60, wrap=True),
            ColumnSpec("Issue", CALCULATED, "Why the filename was flagged.", 40),
        ),
        stats=stats,
        extra_notes=[
            "Gaps on this tab are computed from FILENAMES ONLY. Each PDF may carry "
            "several Bates-stamped pages, so most of these gaps are expected to be "
            "explained by the page counts of the preceding documents. A definitive "
            "gap analysis requires the resolved begin/end ranges produced in Phase 3 "
            "(report 10_Bates_Gap_and_Overlap_Report.xlsx).",
        ],
    )

    # 03 - Exact duplicates
    duplicate_rows: list[dict[str, Any]] = []
    for record in database.iter_files(
        where="exact_dup_group IS NOT NULL",
    ):
        duplicate_rows.append(
            {
                "Exact duplicate group": record["exact_dup_group"],
                "SHA-256": record["sha256"],
                "Filename": record["filename"],
                "Relative path": record["rel_path"],
                "Open file": "Open",
                "File size (bytes)": record["file_size"],
                "Page count": record["page_count"],
                "Bates from filename": record["bates_filename_raw"],
                "_abs_path": record["abs_path"],
            }
        )
    duplicate_rows.sort(key=lambda row: (str(row["Exact duplicate group"]), str(row["Relative path"])))

    workbook = Workbook()
    write_sheet(
        workbook,
        "Exact Duplicates",
        [
            "Exact duplicate group",
            "SHA-256",
            "Filename",
            "Relative path",
            "Open file",
            "File size (bytes)",
            "Page count",
            "Bates from filename",
        ],
        duplicate_rows,
        config=config,
        hyperlink_column="Open file",
    )
    written["03_Exact_Duplicate_Report.xlsx"] = finalise_workbook(
        workbook,
        reports_dir / "03_Exact_Duplicate_Report.xlsx",
        specs=(
            ColumnSpec("Exact duplicate group", CALCULATED, "Group key derived from the shared SHA-256.", 24),
            ColumnSpec("SHA-256", CALCULATED, "Digest shared by every member of the group.", 26),
            ColumnSpec("Filename", OBSERVED, "Base filename.", 34),
            ColumnSpec("Relative path", OBSERVED, "Path relative to the production folder.", 40),
            ColumnSpec("File size (bytes)", OBSERVED, "On-disk size.", 16),
            ColumnSpec("Page count", OBSERVED, "Pages read from the PDF structure.", 12),
            ColumnSpec("Bates from filename", OBSERVED, "Bates value parsed from the filename.", 22),
        ),
        stats=stats,
        extra_notes=[
            "Members of a group are byte-for-byte identical. Producing the same "
            "document at two Bates ranges is normal in a large production and is "
            "not, by itself, an irregularity.",
        ],
    )

    # 04 - Unreadable / corrupt
    problem_rows = [
        {
            "Filename": record["filename"],
            "Relative path": record["rel_path"],
            "Open file": "Open",
            "Condition": (
                "Cloud placeholder (not downloaded)"
                if record["is_cloud_placeholder"]
                else "Password-protected"
                if record["is_encrypted"]
                else "Corrupt or unreadable"
                if record["is_corrupt"]
                else "Processing error"
            ),
            "File size (bytes)": record["file_size"],
            "Error": record["pdf_open_error"] or record["error_message"],
            "Recommended action": (
                "Right-click the file in Finder and choose 'Download now', then re-run Phase 1"
                if record["is_cloud_placeholder"]
                else "Obtain the password from the producing party; this tool never guesses one"
                if record["is_encrypted"]
                else "Request a replacement copy from the producing party"
            ),
            "_abs_path": record["abs_path"],
        }
        for record in database.iter_files(
            where=(
                "is_corrupt = 1 OR is_encrypted = 1 OR is_cloud_placeholder = 1"
                " OR processing_status = 'error'"
            )
        )
    ]

    workbook = Workbook()
    write_sheet(
        workbook,
        "Unreadable or Corrupt",
        [
            "Filename",
            "Relative path",
            "Open file",
            "Condition",
            "File size (bytes)",
            "Error",
            "Recommended action",
        ],
        problem_rows,
        config=config,
        hyperlink_column="Open file",
    )
    written["04_Unreadable_or_Corrupt_Files.xlsx"] = finalise_workbook(
        workbook,
        reports_dir / "04_Unreadable_or_Corrupt_Files.xlsx",
        specs=(
            ColumnSpec("Condition", OBSERVED, "Why the file could not be read.", 30),
            ColumnSpec("Error", OBSERVED, "Verbatim error text from the PDF library or OS.", 50, wrap=True),
            ColumnSpec("Recommended action", INFERRED, "Suggested next step. Cloud placeholders are a sync issue, not a defect in the production.", 60, wrap=True),
        ),
        stats=stats,
        extra_notes=[
            "A cloud placeholder is a file Google Drive has not yet materialised "
            "locally. It is reported separately from corruption because the two "
            "call for completely different follow-up.",
        ],
    )

    written["Phase_1_Summary.md"] = write_phase1_summary(config, database, stats)
    return written


def write_phase1_summary(
    config: Config, database: Database, stats: Mapping[str, Any]
) -> Path:
    """Write the human-readable Phase 1 summary."""
    from src import bates as bates_mod

    path = config.paths.output_folder / "Phase_1_Summary.md"
    total_bytes = int(stats.get("total_bytes", 0) or 0)
    distribution = stats.get("page_count_distribution", {}) or {}

    top_gaps = bates_mod.find_filename_gaps(
        int(row["bates_filename_num"])
        for row in database.query(
            "SELECT bates_filename_num FROM files WHERE bates_filename_num IS NOT NULL"
        )
    )
    top_gaps.sort(key=lambda gap: gap.missing_count, reverse=True)

    lines = [
        "# Phase 1 - Structural Inventory Summary",
        "",
        f"- Generated: {datetime.now().isoformat(timespec='seconds')}",
        f"- Tool: {APP_NAME} v{APP_VERSION}",
        f"- Source folder (read-only): `{config.paths.source_folder}`",
        f"- Derived index folder: `{config.paths.output_folder}`",
        f"- Run id: `{stats.get('run_id', '')}`",
        "",
        "> **Derived metadata notice.** " + DERIVED_NOTICE,
        "",
        "## What Phase 1 did and did not do",
        "",
        "Phase 1 collected structural facts only. It did **not** extract, read, "
        "store, display, or transmit any substantive document text. PDFs were "
        "opened read-only far enough to read a page count and an encryption flag, "
        "then closed. No source file was renamed, moved, altered, annotated, "
        "OCR'd, combined, or deleted.",
        "",
        "## Collection totals",
        "",
        "| Statistic | Value |",
        "| --- | ---: |",
        f"| Total files | {stats.get('total_files', 0):,} |",
        f"| PDFs | {stats.get('total_pdfs', 0):,} |",
        f"| Non-PDF files | {stats.get('total_non_pdfs', 0):,} |",
        f"| Total storage | {total_bytes / (1024 ** 3):.2f} GiB ({total_bytes:,} bytes) |",
        f"| Total pages (PDFs that opened) | {stats.get('total_pages', 0):,} |",
        f"| Files processed this run | {stats.get('processed_now', 0):,} |",
        f"| Files skipped (already complete) | {stats.get('skipped_resumed', 0):,} |",
        "",
        "## Bates numbering (filename-based, preliminary)",
        "",
        "| Statistic | Value |",
        "| --- | ---: |",
        f"| Minimum apparent Bates number | {stats.get('min_bates')} |",
        f"| Maximum apparent Bates number | {stats.get('max_bates')} |",
        f"| Files with a parsed Bates number | {stats.get('bates_parsed', 0):,} |",
        f"| PDFs with no Bates number in the filename | {stats.get('bates_missing', 0):,} |",
        f"| Non-compliant filenames | {stats.get('noncompliant_filenames', 0):,} |",
        f"| Preliminary gap runs | {stats.get('filename_gap_runs', 0):,} |",
        f"| Preliminary missing numbers | {stats.get('filename_gap_missing_numbers', 0):,} |",
        "",
        "**These gaps are preliminary.** Each PDF in this production may contain "
        "several Bates-stamped pages, so a numeric jump between consecutive "
        "filenames is usually explained by the page count of the preceding "
        "document rather than by a missing document. A definitive gap analysis "
        "requires page-level Bates observation, which happens in Phase 3.",
        "",
    ]

    if top_gaps:
        lines.extend(
            [
                "### Largest preliminary gaps",
                "",
                "| After | Before | Missing count |",
                "| --- | --- | ---: |",
            ]
        )
        for gap in top_gaps[:10]:
            lines.append(
                f"| {bates_mod.format_bates(gap.after)} "
                f"| {bates_mod.format_bates(gap.before)} | {gap.missing_count:,} |"
            )
        lines.append("")

    lines.extend(
        [
            "## Duplicates",
            "",
            "| Statistic | Value |",
            "| --- | ---: |",
            f"| Files sharing a basename with another file | {stats.get('duplicate_filename_files', 0):,} |",
            f"| Duplicate-filename groups | {stats.get('duplicate_filename_groups', 0):,} |",
            f"| Files that are byte-identical to another file | {stats.get('exact_duplicate_files', 0):,} |",
            f"| Exact duplicate groups | {stats.get('exact_duplicate_groups', 0):,} |",
            f"| Redundant copies (files beyond one per group) | {stats.get('redundant_duplicate_files', 0):,} |",
            "",
            "## File condition",
            "",
            "| Statistic | Value |",
            "| --- | ---: |",
            f"| Password-protected PDFs | {stats.get('encrypted', 0):,} |",
            f"| Corrupt or unreadable PDFs | {stats.get('corrupt', 0):,} |",
            f"| Cloud placeholders not yet downloaded | {stats.get('cloud_placeholders', 0):,} |",
            f"| Files that errored during processing | {stats.get('errors', 0):,} |",
            "",
        ]
    )

    if distribution:
        lines.extend(["## Page-count distribution", "", "| Pages | Documents |", "| --- | ---: |"])
        for bucket, count in distribution.items():
            lines.append(f"| {bucket} | {count:,} |")
        lines.append("")

    lines.extend(
        [
            "## Reports produced",
            "",
            "| Report | Contents |",
            "| --- | --- |",
            "| `01_File_Inventory.xlsx` / `.csv` | Every file, with hashes, sizes, page counts, and clickable local links. |",
            "| `02_Bates_Structural_Report.xlsx` | Preliminary filename-based Bates gaps and malformed filenames. |",
            "| `03_Exact_Duplicate_Report.xlsx` | Byte-identical file groups. |",
            "| `04_Unreadable_or_Corrupt_Files.xlsx` | Encrypted, corrupt, and not-yet-downloaded files. |",
            "| `wilson_rode_index.sqlite3` | Authoritative processing state, with an FTS5 table ready for Phase 2. |",
            "| `logs/` | Full JSON-lines audit log of every operation. |",
            "",
            "## Limitations",
            "",
        ]
    )
    lines.extend(f"{index}. {note}" for index, note in enumerate(LIMITATIONS, start=1))
    lines.extend(
        [
            "",
            "## Next step",
            "",
            "Phase 1 is complete and the run has stopped as instructed. Phase 2 "
            "(text and page-condition audit on a ~200-document pilot sample) will "
            "not start without explicit approval.",
            "",
        ]
    )

    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text("\n".join(lines), encoding="utf-8")
    LOG.info("Phase 1 summary written", extra={"path": str(path)})
    return path


# ---------------------------------------------------------------------------
# Phase 2 reports
# ---------------------------------------------------------------------------


def write_phase2_reports(
    config: Config,
    database: Database,
    stats: Mapping[str, Any],
    *,
    file_ids: Sequence[int] | None = None,
    pilot: bool = True,
) -> dict[str, Path]:
    """Generate the Phase 2 page-condition deliverables.

    Args:
        config: Loaded configuration.
        database: Processing-state database with pages populated.
        stats: Aggregate Phase 2 statistics.
        file_ids: Restrict the reports to these files (the pilot sample).
        pilot: When True the summary is written as the pilot summary.

    Returns:
        Mapping of report key to written path.
    """
    reports_dir = config.paths.reports_dir
    written: dict[str, Path] = {}
    scope = ""
    params: list[Any] = []
    if file_ids:
        marks = ",".join("?" for _ in file_ids)
        scope = f" AND f.id IN ({marks})"
        params = list(file_ids)

    page_rows = [
        {
            "Filename": row["filename"],
            "Relative path": row["rel_path"],
            "Open file": "Open",
            "Page": row["page_number"],
            "Condition": row["condition"],
            "Blank subtype": row["blank_subtype"],
            "Characters": row["char_count"],
            "Words": row["word_count"],
            "Has text layer": _yes_no(row["has_text_layer"]),
            "Image count": row["image_count"],
            "Image coverage": row["image_coverage"],
            "Bates observed on page": row["bates_observed"],
            "OCR status": row["ocr_status"],
            "Basis": row["notes"],
            "_abs_path": row["abs_path"],
        }
        for row in database.query(
            "SELECT p.*, f.filename, f.rel_path, f.abs_path FROM pages p"
            " JOIN files f ON f.id = p.file_id"
            f" WHERE 1=1{scope} ORDER BY f.rel_path, p.page_number",
            params,
        )
    ]

    workbook = Workbook()
    write_sheet(
        workbook,
        "Page Condition Audit",
        [
            "Filename",
            "Relative path",
            "Open file",
            "Page",
            "Condition",
            "Blank subtype",
            "Characters",
            "Words",
            "Has text layer",
            "Image count",
            "Image coverage",
            "Bates observed on page",
            "OCR status",
            "Basis",
        ],
        page_rows,
        config=config,
        hyperlink_column="Open file",
    )
    written["05_Page_Condition_Audit.xlsx"] = finalise_workbook(
        workbook,
        reports_dir / "05_Page_Condition_Audit.xlsx",
        specs=(
            ColumnSpec("Page", OBSERVED, "1-based page number within the PDF.", 8),
            ColumnSpec("Condition", INFERRED, "Searchable / Image-only / OCR required / OCR completed / Apparent blank / Bates-only page / Corrupt / Uncertain.", 22),
            ColumnSpec("Blank subtype", INFERRED, "For blank-ish pages: truly blank candidate, Bates stamp only, possible separator, possible failed export, or uncertain.", 26),
            ColumnSpec("Characters", OBSERVED, "Characters of text extracted natively from the page.", 12),
            ColumnSpec("Words", OBSERVED, "Whitespace-delimited tokens extracted.", 10),
            ColumnSpec("Has text layer", CALCULATED, "Yes when the character count clears the searchable threshold.", 14),
            ColumnSpec("Image count", OBSERVED, "Embedded images on the page.", 12),
            ColumnSpec("Image coverage", CALCULATED, "Approximate fraction of page area covered by images. Overlapping images inflate this slightly.", 14),
            ColumnSpec("Bates observed on page", OBSERVED, "Bates stamp actually read from the page text - stronger evidence than the filename.", 22),
            ColumnSpec("Basis", CALCULATED, "The measurements that produced the classification.", 60, wrap=True),
        ),
        stats=stats,
    )

    blank_rows = [row for row in page_rows if row["Condition"] in {"Apparent blank", "Bates-only page"}]
    workbook = Workbook()
    write_sheet(
        workbook,
        "Blank and Bates-Only",
        [
            "Filename",
            "Relative path",
            "Open file",
            "Page",
            "Condition",
            "Blank subtype",
            "Characters",
            "Image count",
            "Image coverage",
            "Bates observed on page",
            "Basis",
        ],
        blank_rows,
        config=config,
        hyperlink_column="Open file",
    )
    written["06_Blank_and_Bates_Only_Candidates.xlsx"] = finalise_workbook(
        workbook,
        reports_dir / "06_Blank_and_Bates_Only_Candidates.xlsx",
        specs=(
            ColumnSpec("Blank subtype", INFERRED, "Distinguishes a truly blank candidate from a Bates-stamp-only page, a separator, or a failed export.", 26),
        ),
        stats=stats,
        extra_notes=[
            "A page listed here is a CANDIDATE. Confirming that a page is truly "
            "blank - as opposed to a failed export or an image the extractor "
            "could not see - requires visual inspection.",
        ],
    )

    ocr_rows = [row for row in page_rows if row["Condition"] in {"OCR required", "Image-only"}]
    workbook = Workbook()
    write_sheet(
        workbook,
        "OCR Required",
        [
            "Filename",
            "Relative path",
            "Open file",
            "Page",
            "Condition",
            "Characters",
            "Image count",
            "Image coverage",
            "OCR status",
            "Basis",
        ],
        ocr_rows,
        config=config,
        hyperlink_column="Open file",
    )
    written["07_OCR_Required.xlsx"] = finalise_workbook(
        workbook,
        reports_dir / "07_OCR_Required.xlsx",
        specs=(
            ColumnSpec("OCR status", CALCULATED, "not_required / required / completed / failed / skipped / unavailable.", 16),
        ),
        stats=stats,
        extra_notes=[
            "OCR always writes to a cache copy inside the derived-index folder. "
            "Source PDFs are never OCR'd in place, modified, or replaced.",
        ],
    )

    failure_rows = [
        {
            "Filename": row["filename"],
            "Relative path": row["rel_path"],
            "Open file": "Open",
            "Extraction status": row["extraction_status"],
            "Page count": row["page_count"],
            "Encrypted": _yes_no(row["is_encrypted"]),
            "Corrupt": _yes_no(row["is_corrupt"]),
            "Cloud placeholder": _yes_no(row["is_cloud_placeholder"]),
            "Error": row["error_message"] or row["pdf_open_error"],
            "_abs_path": row["abs_path"],
        }
        for row in database.query(
            "SELECT f.* FROM files f WHERE (f.extraction_status = 'error'"
            " OR f.is_corrupt = 1 OR f.is_encrypted = 1 OR f.is_cloud_placeholder = 1)"
            f"{scope} ORDER BY f.rel_path",
            params,
        )
    ]
    workbook = Workbook()
    write_sheet(
        workbook,
        "Extraction Failures",
        [
            "Filename",
            "Relative path",
            "Open file",
            "Extraction status",
            "Page count",
            "Encrypted",
            "Corrupt",
            "Cloud placeholder",
            "Error",
        ],
        failure_rows,
        config=config,
        hyperlink_column="Open file",
    )
    written["08_Extraction_Failures.xlsx"] = finalise_workbook(
        workbook,
        reports_dir / "08_Extraction_Failures.xlsx",
        specs=(
            ColumnSpec("Extraction status", CALCULATED, "Outcome of the text-extraction attempt.", 18),
            ColumnSpec("Error", OBSERVED, "Verbatim error text.", 60, wrap=True),
        ),
        stats=stats,
    )

    written["Phase_2_Pilot_Summary.md" if pilot else "Phase_2_Summary.md"] = _write_phase2_summary(
        config, stats, pilot=pilot
    )
    return written


def _write_phase2_summary(config: Config, stats: Mapping[str, Any], *, pilot: bool) -> Path:
    """Write the Phase 2 summary document."""
    name = "Phase_2_Pilot_Summary.md" if pilot else "Phase_2_Summary.md"
    path = config.paths.output_folder / name
    scope = "pilot sample" if pilot else "full production"

    lines = [
        f"# Phase 2 - Text and Page-Condition Audit ({scope})",
        "",
        f"- Generated: {datetime.now().isoformat(timespec='seconds')}",
        f"- Tool: {APP_NAME} v{APP_VERSION}",
        f"- Run id: `{stats.get('run_id', '')}`",
        "",
        "> **Derived metadata notice.** " + DERIVED_NOTICE,
        "",
        "## Scope",
        "",
        f"| Statistic | Value |",
        "| --- | ---: |",
        f"| Documents examined | {stats.get('documents', 0):,} |",
        f"| Pages examined | {stats.get('pages', 0):,} |",
        f"| Documents that failed extraction | {stats.get('failed', 0):,} |",
        "",
        "## Page conditions",
        "",
        "| Condition | Pages |",
        "| --- | ---: |",
    ]
    for condition, count in (stats.get("page_conditions", {}) or {}).items():
        lines.append(f"| {condition} | {count:,} |")

    lines.extend(
        [
            "",
            "## Blank and Bates-only breakdown",
            "",
            "| Subtype | Pages |",
            "| --- | ---: |",
        ]
    )
    for subtype, count in (stats.get("blank_subtypes", {}) or {}).items():
        lines.append(f"| {subtype} | {count:,} |")

    lines.extend(
        [
            "",
            "## Bates observation",
            "",
            "| Statistic | Value |",
            "| --- | ---: |",
            f"| Pages with a Bates stamp read from the page | {stats.get('pages_with_observed_bates', 0):,} |",
            f"| Documents where observed Bates matched the filename | {stats.get('bates_filename_agreement', 0):,} |",
            f"| Documents where observed Bates DISAGREED with the filename | {stats.get('bates_filename_disagreement', 0):,} |",
            "",
            "Filename Bates values are not trusted on their own. Where a stamp was "
            "readable on the page, the page value governs.",
            "",
            "## OCR",
            "",
            "| Statistic | Value |",
            "| --- | ---: |",
            f"| Pages needing OCR | {stats.get('ocr_required_pages', 0):,} |",
            f"| Pages OCR completed | {stats.get('ocr_completed_pages', 0):,} |",
            f"| OCR failures | {stats.get('ocr_failures', 0):,} |",
            f"| OCR available on this machine | {stats.get('ocr_available', 'unknown')} |",
            "",
            "## Accuracy problems and known limitations",
            "",
        ]
    )
    problems = stats.get("accuracy_notes") or [
        "No systematic accuracy problems were detected automatically. Automated "
        "detection cannot substitute for the manual QC sample described in the README.",
    ]
    lines.extend(f"- {note}" for note in problems)
    lines.extend(["", "## Limitations", ""])
    lines.extend(f"{index}. {note}" for index, note in enumerate(LIMITATIONS, start=1))

    if pilot:
        lines.extend(
            [
                "",
                "## Next step",
                "",
                "The pilot has stopped as instructed. Review the accuracy notes above "
                "and the QC sample before approving a full-production Phase 2/3 run.",
                "",
            ]
        )

    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text("\n".join(lines), encoding="utf-8")
    LOG.info("Phase 2 summary written", extra={"path": str(path)})
    return path


# ---------------------------------------------------------------------------
# Phase 3 reports
# ---------------------------------------------------------------------------

MASTER_COLUMNS: tuple[ColumnSpec, ...] = (
    ColumnSpec("Begin Bates", OBSERVED, "First Bates number of the document. Observed on the page where readable, otherwise derived - see 'Bates source'.", 18),
    ColumnSpec("End Bates", CALCULATED, "Last Bates number of the document. Frequently calculated from the begin value plus the page count.", 18),
    ColumnSpec("Bates confidence", CALCULATED, "Confirmed / Probable / Possible / Unknown.", 16),
    ColumnSpec("Bates source", CALCULATED, "Observed on page / Parsed from filename / Calculated / Inferred.", 20),
    ColumnSpec("Filename", OBSERVED, "Base filename as produced.", 34),
    ColumnSpec("Relative path", OBSERVED, "Path relative to the production folder.", 40),
    ColumnSpec("Open file", CALCULATED, "Clickable file:// link to the local source PDF.", 12),
    ColumnSpec("Page count", OBSERVED, "Pages read from the PDF structure.", 11),
    ColumnSpec("File size (bytes)", OBSERVED, "On-disk size.", 15),
    ColumnSpec("File hash", CALCULATED, "SHA-256 of the file bytes.", 26),
    ColumnSpec("Document date", INFERRED, "Apparent date of the document. Taken from an email header where present, otherwise the earliest plausible date in the leading text.", 15),
    ColumnSpec("Date confidence", CALCULATED, "Confirmed / Probable / Possible / Unknown.", 15),
    ColumnSpec("Author or From", INFERRED, "Email sender where present, otherwise a signature-block name.", 30),
    ColumnSpec("To", OBSERVED, "Top-level message recipients, semicolon-delimited. Quoted headers from earlier messages in a chain are excluded.", 34),
    ColumnSpec("CC", OBSERVED, "Top-level message CC recipients.", 30),
    ColumnSpec("BCC", OBSERVED, "Top-level message BCC recipients. Rarely present in a production copy.", 24),
    ColumnSpec("Subject or Title", INFERRED, "Email subject where present, otherwise the first substantial line.", 44),
    ColumnSpec("Document type", INFERRED, "Deterministic keyword classification. Always read with the confidence column.", 22),
    ColumnSpec("Document type confidence", CALCULATED, "Winning type's share of total classification score, 0-1.", 20),
    ColumnSpec("Court", INFERRED, "Court name found in the caption area.", 34),
    ColumnSpec("Case number", INFERRED, "Docket or case number found in the caption area.", 20),
    ColumnSpec("Parent Bates", INFERRED, "Begin Bates of a document inferred to be this one's parent email. NOT native family metadata.", 18),
    ColumnSpec("Attachment Bates", INFERRED, "Begin Bates values of documents inferred to be attachments to this one. NOT native family metadata.", 24),
    ColumnSpec("Family confidence", CALCULATED, "Weakest confidence among this document's inferred relationships.", 17),
    ColumnSpec("Exact duplicate group", CALCULATED, "Group of byte-identical files (SHA-256).", 21),
    ColumnSpec("Text duplicate group", CALCULATED, "Group whose normalised text is identical - the same document rendered differently.", 20),
    ColumnSpec("Near duplicate group", CALCULATED, "SimHash/LSH group of highly similar documents.", 20),
    ColumnSpec("Issue tags", INFERRED, "Screening tags whose configured terms appeared in extracted text. NOT legal conclusions.", 34, wrap=True),
    ColumnSpec("Priority score", CALCULATED, "Additive score from the configured factors. Means 'review early'; does NOT indicate malpractice or liability.", 13),
    ColumnSpec("Priority score explanation", CALCULATED, "Every factor that contributed, with its points.", 60, wrap=True),
    ColumnSpec("Searchable status", CALCULATED, "Fully searchable / Partially searchable / No searchable text / Unknown.", 18),
    ColumnSpec("OCR status", CALCULATED, "Whether OCR was needed, run, or unavailable.", 14),
    ColumnSpec("Apparent blank status", INFERRED, "Whether the document appears to be blank or a separator.", 20),
    ColumnSpec("Extraction confidence", CALCULATED, "Overall trust in this row's derived fields.", 19),
    ColumnSpec("Manual review required", INFERRED, "Yes when confidence is low or the document could not be classified.", 20),
    ColumnSpec("Processing notes", CALCULATED, "Why fields were assigned or left blank.", 60, wrap=True),
)

MASTER_COLUMN_NAMES: tuple[str, ...] = tuple(spec.name for spec in MASTER_COLUMNS)


def write_phase3_reports(
    config: Config,
    database: Database,
    stats: Mapping[str, Any],
    *,
    master_rows: Sequence[Mapping[str, Any]],
    gap_rows: Sequence[Mapping[str, Any]],
    overlap_rows: Sequence[Mapping[str, Any]],
    duplicate_rows: Sequence[Mapping[str, Any]],
    family_rows: Sequence[Mapping[str, Any]],
    issue_rows: Sequence[Mapping[str, Any]],
    blank_rows: Sequence[Mapping[str, Any]],
    error_rows: Sequence[Mapping[str, Any]],
) -> dict[str, Path]:
    """Generate every Phase 3 deliverable.

    All row sets are supplied by :mod:`src.cli`, which owns the database
    queries. This function is presentation only.
    """
    reports_dir = config.paths.reports_dir
    written: dict[str, Path] = {}

    # 09 - master index
    workbook = Workbook()
    write_sheet(
        workbook,
        "Derived Document Index",
        MASTER_COLUMN_NAMES,
        master_rows,
        config=config,
        hyperlink_column="Open file",
    )
    written["09_Derived_Document_Index.xlsx"] = finalise_workbook(
        workbook,
        reports_dir / "09_Derived_Document_Index.xlsx",
        specs=MASTER_COLUMNS,
        stats=stats,
    )
    written["09_Derived_Document_Index.csv"] = write_csv(
        reports_dir / "09_Derived_Document_Index.csv", MASTER_COLUMN_NAMES, master_rows
    )

    # 10 - Bates gaps and overlaps
    workbook = Workbook()
    write_sheet(
        workbook,
        "Gaps",
        ["Gap after Bates", "Gap before Bates", "Missing numbers", "Count missing", "Basis"],
        gap_rows,
        config=config,
    )
    write_sheet(
        workbook,
        "Overlaps",
        ["Document A", "A range", "Document B", "B range", "Overlap", "Note"],
        overlap_rows,
        config=config,
    )
    written["10_Bates_Gap_and_Overlap_Report.xlsx"] = finalise_workbook(
        workbook,
        reports_dir / "10_Bates_Gap_and_Overlap_Report.xlsx",
        specs=(
            ColumnSpec("Missing numbers", CALCULATED, "Bates numbers no document's resolved range covers.", 22),
            ColumnSpec("Basis", CALCULATED, "Computed from resolved begin/end ranges with page counts accounted for - not from filenames alone.", 60, wrap=True),
            ColumnSpec("Overlap", CALCULATED, "Bates numbers claimed by more than one document.", 22),
        ),
        stats=stats,
        extra_notes=[
            "This report supersedes the preliminary filename-based gap analysis in "
            "02_Bates_Structural_Report.xlsx. Ranges here account for page counts "
            "and, where readable, Bates stamps observed on the pages themselves.",
            "An overlap usually means two documents claim the same Bates numbers - "
            "commonly a re-produced copy. It can also indicate a resolution error "
            "where a page count was wrong; both are worth a look.",
        ],
    )

    # 11 - blanks and extraction failures
    workbook = Workbook()
    write_sheet(
        workbook,
        "Blank and Failure Candidates",
        [
            "Begin Bates",
            "Filename",
            "Relative path",
            "Open file",
            "Page count",
            "Condition",
            "Detail",
            "Recommended action",
        ],
        blank_rows,
        config=config,
        hyperlink_column="Open file",
    )
    written["11_Blank_and_Extraction_Failure_Report.xlsx"] = finalise_workbook(
        workbook,
        reports_dir / "11_Blank_and_Extraction_Failure_Report.xlsx",
        specs=(
            ColumnSpec("Condition", INFERRED, "Blank candidate, Bates-only, extraction failure, or unreadable.", 24),
            ColumnSpec("Detail", CALCULATED, "The measurements behind the classification.", 50, wrap=True),
        ),
        stats=stats,
    )

    # 12 - duplicate groups
    workbook = Workbook()
    write_sheet(
        workbook,
        "Duplicate Groups",
        [
            "Group",
            "Method",
            "Relation",
            "Similarity",
            "Primary",
            "Begin Bates",
            "Filename",
            "Relative path",
            "Open file",
            "Page count",
        ],
        duplicate_rows,
        config=config,
        hyperlink_column="Open file",
    )
    written["12_Duplicate_Groups.xlsx"] = finalise_workbook(
        workbook,
        reports_dir / "12_Duplicate_Groups.xlsx",
        specs=(
            ColumnSpec("Method", CALCULATED, "exact_binary (SHA-256), normalized_text (identical normalised text), or near_duplicate (SimHash/LSH).", 18),
            ColumnSpec("Relation", INFERRED, "Exact binary duplicate / Same text, different PDF rendering / Email chain containing an earlier email / Draft and final / Near duplicate / Possible duplicate.", 40),
            ColumnSpec("Similarity", CALCULATED, "1.0 for exact methods; for near duplicates, 1 minus the normalised Hamming distance.", 12),
            ColumnSpec("Primary", CALCULATED, "Yes for the stable representative of the group.", 10),
        ),
        stats=stats,
        extra_notes=[
            "Near-duplicate detection uses banded LSH over 64-bit SimHash values. "
            "It does not perform an O(n^2) comparison across the production, so "
            "recall is high but not guaranteed exhaustive.",
        ],
    )

    # 13 - family inferences
    workbook = Workbook()
    write_sheet(
        workbook,
        "Family Inferences",
        [
            "Parent Bates",
            "Child Bates",
            "Relationship type",
            "Confidence",
            "Reason",
            "Parent filename",
            "Child filename",
            "Open parent",
        ],
        family_rows,
        config=config,
        hyperlink_column="Open parent",
    )
    written["13_Document_Family_Inferences.xlsx"] = finalise_workbook(
        workbook,
        reports_dir / "13_Document_Family_Inferences.xlsx",
        specs=(
            ColumnSpec("Relationship type", INFERRED, "Parent email to attachment, or a sequential continuation.", 28),
            ColumnSpec("Confidence", CALCULATED, "Confirmed / Probable / Possible / Unknown.", 13),
            ColumnSpec("Reason", INFERRED, "Every signal that supported the inference.", 80, wrap=True),
        ),
        stats=stats,
        extra_notes=[
            "EVERY RELATIONSHIP ON THIS TAB IS AN INFERENCE. This production "
            "included no load file, so native parent/child family metadata does "
            "not exist here. Relationships were reconstructed from Bates "
            "adjacency, attachment names, dates, and subjects. Verify before "
            "relying on any of them.",
        ],
    )

    # 14 - issue tag index
    workbook = Workbook()
    write_sheet(
        workbook,
        "Issue Tag Index",
        [
            "Issue tag",
            "Matching term",
            "Hits",
            "Page",
            "Locator",
            "Begin Bates",
            "Filename",
            "Relative path",
            "Open file",
        ],
        issue_rows,
        config=config,
        hyperlink_column="Open file",
    )
    written["14_Issue_Tag_Index.xlsx"] = finalise_workbook(
        workbook,
        reports_dir / "14_Issue_Tag_Index.xlsx",
        specs=(
            ColumnSpec("Issue tag", INFERRED, "Configured screening category. NOT a legal conclusion.", 24),
            ColumnSpec("Matching term", OBSERVED, "The configured term that matched.", 28),
            ColumnSpec("Hits", CALCULATED, "Total matches for this term in this document.", 8),
            ColumnSpec("Page", OBSERVED, "Page the first match appeared on.", 8),
            ColumnSpec("Locator", CALCULATED, "Structural pointer (page and character offset). Deliberately not an excerpt, so confidential passages stay out of this workbook.", 30),
        ),
        stats=stats,
        extra_notes=[
            "Issue tags are a screening tool. A hit means a configured term "
            "appeared in extracted text - nothing more. Terms are intentionally "
            "over-inclusive so that relevant documents are not missed.",
            "Locators are structural by design. To include surrounding text, raise "
            "issue_tags.context_chars in config.yaml - but note that doing so puts "
            "document content into a spreadsheet.",
        ],
    )

    # 15-20 - review queues
    queue_specs = (
        ("15_Priority_Review_Queue.xlsx", "Priority Review Queue", None,
         "Documents at or above the configured priority threshold."),
        ("16_John_Wilson_Review_Queue.xlsx", "John Wilson Queue", "JOHN_WILSON",
         "Documents tagged JOHN_WILSON."),
        ("17_Statute_of_Frauds_Review_Queue.xlsx", "Statute of Frauds Queue", "STATUTE_OF_FRAUDS",
         "Documents tagged STATUTE_OF_FRAUDS."),
        ("18_Expert_and_Damages_Review_Queue.xlsx", "Expert and Damages Queue", ("EXPERT_FAILURE", "DAMAGES"),
         "Documents tagged EXPERT_FAILURE or DAMAGES."),
        ("19_Client_File_Review_Queue.xlsx", "Client File Queue", "CLIENT_FILE",
         "Documents tagged CLIENT_FILE."),
        ("20_Insurance_Review_Queue.xlsx", "Insurance Queue", "INSURANCE",
         "Documents tagged INSURANCE."),
    )
    threshold = float(config.get("priority.queue_threshold", 20) or 0)
    max_rows = int(config.get("priority.queue_max_rows", 0) or 0)

    for filename, sheet_title, tag_filter, description in queue_specs:
        if tag_filter is None:
            selected = [
                row for row in master_rows
                if float(row.get("Priority score") or 0) >= threshold
            ]
        else:
            wanted = (tag_filter,) if isinstance(tag_filter, str) else tag_filter
            selected = [
                row for row in master_rows
                if any(tag in str(row.get("Issue tags") or "") for tag in wanted)
            ]
        selected = sorted(
            selected, key=lambda row: float(row.get("Priority score") or 0), reverse=True
        )
        if max_rows:
            selected = selected[:max_rows]

        workbook = Workbook()
        write_sheet(
            workbook,
            sheet_title,
            MASTER_COLUMN_NAMES,
            selected,
            config=config,
            hyperlink_column="Open file",
        )
        written[filename] = finalise_workbook(
            workbook,
            reports_dir / filename,
            specs=MASTER_COLUMNS,
            stats={**stats, "queue": description, "queue_rows": len(selected)},
            extra_notes=[
                f"Queue contents: {description}",
                "Queue membership is a triage decision, not a relevance, privilege, "
                "or admissibility determination.",
            ],
        )

    # Processing errors
    workbook = Workbook()
    write_sheet(
        workbook,
        "Processing Errors",
        ["Filename", "Relative path", "Open file", "Stage", "Status", "Error", "Recommended action"],
        error_rows,
        config=config,
        hyperlink_column="Open file",
    )
    written["Processing_Errors.xlsx"] = finalise_workbook(
        workbook,
        reports_dir / "Processing_Errors.xlsx",
        specs=(
            ColumnSpec("Stage", CALCULATED, "Which phase recorded the error.", 14),
            ColumnSpec("Error", OBSERVED, "Verbatim error text.", 60, wrap=True),
        ),
        stats=stats,
    )

    written["Final_Run_Summary.md"] = _write_final_summary(config, stats)
    return written


def _write_final_summary(config: Config, stats: Mapping[str, Any]) -> Path:
    """Write the Phase 3 final run summary."""
    path = config.paths.output_folder / "Final_Run_Summary.md"
    lines = [
        "# Final Run Summary",
        "",
        f"- Generated: {datetime.now().isoformat(timespec='seconds')}",
        f"- Tool: {APP_NAME} v{APP_VERSION}",
        f"- Run id: `{stats.get('run_id', '')}`",
        f"- Source folder (read-only): `{config.paths.source_folder}`",
        f"- Derived index folder: `{config.paths.output_folder}`",
        "",
        "> **Derived metadata notice.** " + DERIVED_NOTICE,
        "",
        "## Totals",
        "",
        "| Statistic | Value |",
        "| --- | ---: |",
    ]
    for key, value in stats.items():
        if isinstance(value, (str, int, float)) and not isinstance(value, bool):
            lines.append(f"| {key.replace('_', ' ')} | {value} |")

    for section_key, heading in (
        ("document_types", "Document types"),
        ("bates_confidence", "Bates confidence"),
        ("issue_tag_counts", "Issue tag counts"),
        ("duplicate_counts", "Duplicate groups by method"),
        ("family_confidence", "Family inference confidence"),
    ):
        section = stats.get(section_key) or {}
        if not section:
            continue
        lines.extend(["", f"## {heading}", "", "| Value | Count |", "| --- | ---: |"])
        for key, value in section.items():
            lines.append(f"| {key} | {value:,} |")

    lines.extend(
        [
            "",
            "## Reports produced",
            "",
            "See `reports/` for workbooks 09 through 20, `Processing_Errors.xlsx`, "
            "and the CSV copy of the master index. The SQLite database carries the "
            "authoritative state and an FTS5 table for full-text search.",
            "",
            "## Accuracy",
            "",
            "No perfect-accuracy claim is made. Measured pilot accuracy is recorded "
            "in `Phase_2_Pilot_Summary.md`. Review the QC sample described in the "
            "README before relying on any derived field.",
            "",
            "## Limitations",
            "",
        ]
    )
    lines.extend(f"{index}. {note}" for index, note in enumerate(LIMITATIONS, start=1))
    lines.append("")

    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text("\n".join(lines), encoding="utf-8")
    LOG.info("Final run summary written", extra={"path": str(path)})
    return path


__all__ = [
    "CALCULATED",
    "ColumnSpec",
    "DERIVED_NOTICE",
    "INFERRED",
    "LIMITATIONS",
    "MASTER_COLUMNS",
    "MASTER_COLUMN_NAMES",
    "OBSERVED",
    "add_data_dictionary",
    "add_methodology",
    "add_run_statistics",
    "finalise_workbook",
    "write_csv",
    "write_phase1_reports",
    "write_phase1_summary",
    "write_phase2_reports",
    "write_phase3_reports",
    "write_sheet",
]
