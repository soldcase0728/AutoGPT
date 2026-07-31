"""Integration tests: Phase 1, resumability, and read-only source protection.

The read-only tests are the most important in the suite. They fingerprint
every source file before a run and assert byte-for-byte equality
afterwards, so any code path that renames, rewrites, OCRs in place, or
deletes a source document fails loudly.
"""

from __future__ import annotations

import hashlib
from pathlib import Path

import pytest
import yaml

from src.config import ConfigError, load_config
from src.database import STATUS_DONE, STATUS_PLACEHOLDER, Database
from src.inventory import probe_pdf_structure, run_inventory, sha256_file, should_ignore, walk_source
from src.pdf_extract import ocr_cache_path, run_ocr
from src.reports import write_phase1_reports


def fingerprint_tree(root: Path) -> dict[str, tuple[int, str, float]]:
    """Capture size, digest, and mtime for every file beneath ``root``."""
    snapshot: dict[str, tuple[int, str, float]] = {}
    for path in sorted(root.rglob("*")):
        if path.is_file():
            data = path.read_bytes()
            snapshot[str(path.relative_to(root))] = (
                len(data),
                hashlib.sha256(data).hexdigest(),
                path.stat().st_mtime,
            )
    return snapshot


class TestReadOnlySourceProtection:
    """The source production must never be touched."""

    def test_phase1_leaves_source_byte_identical(self, config, database, source_tree: Path) -> None:
        before = fingerprint_tree(source_tree)
        run_inventory(config, database, progress=False)
        after = fingerprint_tree(source_tree)
        assert before == after, "Phase 1 modified the source folder"

    def test_report_generation_leaves_source_untouched(
        self, config, database, source_tree: Path
    ) -> None:
        stats = run_inventory(config, database, progress=False)
        before = fingerprint_tree(source_tree)
        write_phase1_reports(config, database, stats.as_dict())
        assert fingerprint_tree(source_tree) == before

    def test_no_files_created_in_source(self, config, database, source_tree: Path) -> None:
        names_before = {str(p.relative_to(source_tree)) for p in source_tree.rglob("*")}
        stats = run_inventory(config, database, progress=False)
        write_phase1_reports(config, database, stats.as_dict())
        names_after = {str(p.relative_to(source_tree)) for p in source_tree.rglob("*")}
        assert names_before == names_after

    def test_all_output_lands_outside_the_source(self, config, database) -> None:
        stats = run_inventory(config, database, progress=False)
        written = write_phase1_reports(config, database, stats.as_dict())
        for path in written.values():
            assert config.paths.source_folder not in path.parents

    def test_config_rejects_output_inside_source(self, tmp_path: Path, source_tree: Path) -> None:
        template = yaml.safe_load(
            (Path(__file__).resolve().parent.parent / "config.yaml").read_text(encoding="utf-8")
        )
        template["paths"]["source_folder"] = str(source_tree)
        template["paths"]["output_folder"] = str(source_tree / "derived")
        path = tmp_path / "bad.yaml"
        path.write_text(yaml.safe_dump(template), encoding="utf-8")
        with pytest.raises(ConfigError, match="inside"):
            load_config(path).validate_paths()

    def test_config_rejects_source_inside_output(self, tmp_path: Path, source_tree: Path) -> None:
        template = yaml.safe_load(
            (Path(__file__).resolve().parent.parent / "config.yaml").read_text(encoding="utf-8")
        )
        template["paths"]["source_folder"] = str(source_tree)
        template["paths"]["output_folder"] = str(source_tree.parent)
        path = tmp_path / "bad2.yaml"
        path.write_text(yaml.safe_dump(template), encoding="utf-8")
        with pytest.raises(ConfigError):
            load_config(path).validate_paths()

    def test_ocr_refuses_to_write_over_its_input(self, source_tree: Path) -> None:
        pdf = source_tree / "WILSONRODE000001-null.pdf"
        with pytest.raises(ValueError, match="in place"):
            run_ocr(pdf, pdf)

    def test_ocr_cache_path_is_inside_the_cache_dir(self, config) -> None:
        target = ocr_cache_path(config.paths.ocr_cache_dir, "a/b.pdf", "a" * 64)
        assert config.paths.ocr_cache_dir in target.parents
        assert config.paths.source_folder not in target.parents


class TestWalking:
    """Discovery and ignore rules."""

    def test_ignores_mac_and_drive_noise(self) -> None:
        patterns = [".DS_Store", "._*", "desktop.ini"]
        assert should_ignore(".DS_Store", patterns)
        assert should_ignore("._WILSONRODE000001-null.pdf", patterns)
        assert not should_ignore("WILSONRODE000001-null.pdf", patterns)

    def test_walk_skips_ignored_names(self, source_tree: Path) -> None:
        found = {
            path.name
            for path in walk_source(source_tree, ignore_names=[".DS_Store"], ignore_dirs=[])
        }
        assert ".DS_Store" not in found
        assert "WILSONRODE000001-null.pdf" in found

    def test_walk_skips_ignored_directories(self, source_tree: Path) -> None:
        nested = source_tree / ".Trash"
        nested.mkdir()
        (nested / "junk.pdf").write_bytes(b"%PDF-1.4 junk")
        found = {
            path.name
            for path in walk_source(source_tree, ignore_names=[], ignore_dirs=[".Trash"])
        }
        assert "junk.pdf" not in found


class TestStructuralProbe:
    """Page-count and condition detection."""

    def test_page_count_read_correctly(self, source_tree: Path) -> None:
        result = probe_pdf_structure(source_tree / "WILSONRODE000002-null.pdf")
        assert result["page_count"] == 2
        assert result["is_corrupt"] is False

    def test_single_page_document(self, source_tree: Path) -> None:
        result = probe_pdf_structure(source_tree / "WILSONRODE000001-null.pdf")
        assert result["page_count"] == 1

    def test_encrypted_document_detected_without_guessing_a_password(
        self, source_tree: Path
    ) -> None:
        result = probe_pdf_structure(source_tree / "WILSONRODE000020-null.pdf")
        assert result["is_encrypted"] is True
        assert result["page_count"] is None
        assert "Password-protected" in result["pdf_open_error"]

    def test_corrupt_file_reported_not_raised(self, tmp_path: Path) -> None:
        broken = tmp_path / "broken.pdf"
        broken.write_bytes(b"%PDF-1.4 this is not really a pdf at all")
        result = probe_pdf_structure(broken)
        assert result["is_corrupt"] is True
        assert result["page_count"] is None

    def test_sha256_matches_hashlib(self, source_tree: Path) -> None:
        path = source_tree / "WILSONRODE000001-null.pdf"
        assert sha256_file(path) == hashlib.sha256(path.read_bytes()).hexdigest()


class TestPhase1Inventory:
    """End-to-end Phase 1 behaviour on the synthetic production."""

    def test_inventory_counts(self, config, database) -> None:
        stats = run_inventory(config, database, progress=False)
        # 8 PDFs (one an encrypted document, one a placeholder stub) plus
        # notes.txt. .DS_Store is ignored by configuration.
        assert stats.total_files == 9
        assert stats.total_pdfs == 8
        assert stats.total_non_pdfs == 1

    def test_bates_range_from_filenames(self, config, database) -> None:
        stats = run_inventory(config, database, progress=False)
        assert stats.min_bates == 1
        assert stats.max_bates == 30

    def test_exact_duplicate_detected(self, config, database) -> None:
        stats = run_inventory(config, database, progress=False)
        assert stats.exact_duplicate_groups == 1
        assert stats.exact_duplicate_files == 2
        assert stats.redundant_duplicate_files == 1

    def test_malformed_filename_flagged(self, config, database) -> None:
        run_inventory(config, database, progress=False)
        row = database.get_file("scan_0012.pdf")
        assert row["bates_filename_num"] is None
        assert row["filename_compliant"] == 0

    def test_encrypted_flagged(self, config, database) -> None:
        run_inventory(config, database, progress=False)
        row = database.get_file("WILSONRODE000020-null.pdf")
        assert row["is_encrypted"] == 1

    def test_cloud_placeholder_not_called_corrupt(self, config, database) -> None:
        run_inventory(config, database, progress=False)
        row = database.get_file("WILSONRODE000030-null.pdf")
        assert row["is_cloud_placeholder"] == 1
        assert row["is_corrupt"] == 0
        assert row["processing_status"] == STATUS_PLACEHOLDER

    def test_filename_gaps_are_preliminary(self, config, database) -> None:
        stats = run_inventory(config, database, progress=False)
        # Filenames jump 1,2,4,10,11,20,30 - several apparent gaps.
        assert stats.filename_gap_runs > 0
        assert stats.filename_gap_missing_numbers > 0

    def test_page_count_distribution_populated(self, config, database) -> None:
        stats = run_inventory(config, database, progress=False)
        assert sum(stats.page_count_distribution.values()) == stats.pdfs_with_page_count

    def test_reports_written_with_expected_names(self, config, database) -> None:
        stats = run_inventory(config, database, progress=False)
        written = write_phase1_reports(config, database, stats.as_dict())
        for name in (
            "01_File_Inventory.xlsx",
            "01_File_Inventory.csv",
            "02_Bates_Structural_Report.xlsx",
            "03_Exact_Duplicate_Report.xlsx",
            "04_Unreadable_or_Corrupt_Files.xlsx",
            "Phase_1_Summary.md",
        ):
            assert name in written
            assert written[name].exists()

    def test_summary_flags_gaps_as_preliminary(self, config, database) -> None:
        stats = run_inventory(config, database, progress=False)
        written = write_phase1_reports(config, database, stats.as_dict())
        text = written["Phase_1_Summary.md"].read_text(encoding="utf-8")
        assert "preliminary" in text.lower()
        assert "DERIVED REVIEW METADATA" in text

    def test_workbooks_have_required_tabs(self, config, database) -> None:
        from openpyxl import load_workbook

        stats = run_inventory(config, database, progress=False)
        written = write_phase1_reports(config, database, stats.as_dict())
        workbook = load_workbook(written["01_File_Inventory.xlsx"])
        assert "Data Dictionary" in workbook.sheetnames
        assert "Methodology and Limitations" in workbook.sheetnames
        assert "Run Statistics" in workbook.sheetnames
        sheet = workbook["File Inventory"]
        assert sheet.freeze_panes == "A2"
        assert sheet.auto_filter.ref is not None

    def test_audit_log_records_the_run(self, config, database) -> None:
        run_inventory(config, database, progress=False)
        operations = {
            row["operation"] for row in database.query("SELECT operation FROM audit_log")
        }
        assert {"phase1_start", "walk_complete", "phase1_complete"} <= operations


class TestResumability:
    """A second run must not redo completed work.

    One file in the synthetic production is a cloud placeholder. Those are
    deliberately retried on every run - they are expected to become
    readable once Drive materialises them - so the "still to do" count
    never drops below one here.
    """

    PLACEHOLDER_RETRIES = 1

    def test_second_run_skips_completed_files(self, config, database) -> None:
        first = run_inventory(config, database, progress=False)
        assert first.processed_now > 0
        second = run_inventory(config, database, progress=False)
        assert second.processed_now == self.PLACEHOLDER_RETRIES
        assert second.skipped_resumed == first.processed_now - self.PLACEHOLDER_RETRIES

    def test_placeholder_is_always_retried(self, config, database) -> None:
        run_inventory(config, database, progress=False)
        run_inventory(config, database, progress=False)
        row = database.get_file("WILSONRODE000030-null.pdf")
        assert row["processing_status"] == STATUS_PLACEHOLDER
        assert row["is_corrupt"] == 0

    def test_totals_are_stable_across_runs(self, config, database) -> None:
        first = run_inventory(config, database, progress=False)
        second = run_inventory(config, database, progress=False)
        assert second.total_files == first.total_files
        assert second.exact_duplicate_groups == first.exact_duplicate_groups

    def test_changed_file_is_reprocessed(self, config, database, source_tree: Path) -> None:
        run_inventory(config, database, progress=False)
        from tests.conftest import make_pdf

        make_pdf(source_tree / "WILSONRODE000004-null.pdf", ["changed content", "second page"])
        third = run_inventory(config, database, progress=False)
        assert third.processed_now == 1 + self.PLACEHOLDER_RETRIES
        assert database.get_file("WILSONRODE000004-null.pdf")["page_count"] == 2

    def test_unchanged_file_keeps_its_recorded_state(self, config, database) -> None:
        run_inventory(config, database, progress=False)
        before = database.get_file("WILSONRODE000001-null.pdf")["last_processed_ts"]
        run_inventory(config, database, progress=False)
        assert database.get_file("WILSONRODE000001-null.pdf")["last_processed_ts"] == before

    def test_new_file_is_picked_up(self, config, database, source_tree: Path) -> None:
        run_inventory(config, database, progress=False)
        from tests.conftest import make_pdf

        make_pdf(source_tree / "WILSONRODE000050-null.pdf", ["new document"])
        result = run_inventory(config, database, progress=False)
        assert result.processed_now == 1 + self.PLACEHOLDER_RETRIES
        assert result.total_files == 10

    def test_force_reprocesses_everything(self, config, database) -> None:
        first = run_inventory(config, database, progress=False)
        forced = run_inventory(config, database, progress=False, force=True)
        assert forced.processed_now == first.processed_now

    def test_needs_processing_detects_version_change(self, config, database) -> None:
        run_inventory(config, database, progress=False)
        row = database.get_file("WILSONRODE000001-null.pdf")
        assert not database.needs_processing(
            "WILSONRODE000001-null.pdf",
            size=int(row["file_size"]),
            mtime_iso=str(row["modified_ts"]),
        )
        assert database.needs_processing(
            "WILSONRODE000001-null.pdf",
            size=int(row["file_size"]),
            mtime_iso=str(row["modified_ts"]),
            app_version="99.0.0",
        )

    def test_unknown_file_always_needs_processing(self, database: Database) -> None:
        assert database.needs_processing("never/seen.pdf", size=1, mtime_iso="2020-01-01")
