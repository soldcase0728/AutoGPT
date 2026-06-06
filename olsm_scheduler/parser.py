"""Parse OLSM input files into data models."""

from __future__ import annotations

import csv
import math
import openpyxl
from collections import Counter, defaultdict
from pathlib import Path
from typing import Optional

from .models import (
    Course, CourseFlag, Constraint, Gender,
    Period, Room, ScheduleData, Student, Teacher, ALL_PERIODS,
)


def _str(val) -> str:
    if val is None:
        return ""
    return str(val).strip()


def _int(val, default: int = 0) -> int:
    if val is None:
        return default
    try:
        return int(val)
    except (ValueError, TypeError):
        return default


def _parse_gender(val: str) -> Optional[Gender]:
    val = val.strip().lower()
    if val in ("g", "girls", "girl", "f", "female"):
        return Gender.GIRLS
    if val in ("b", "boys", "boy", "m", "male"):
        return Gender.BOYS
    return None


def _parse_grade(val) -> int:
    s = _str(val).lower().replace("th", "").replace("grade", "").replace("nd", "").replace("rd", "").replace("st", "").strip()
    for num in ("12", "11", "10", "9"):
        if num in s:
            return int(num)
    try:
        return int(s)
    except ValueError:
        return 0


def _parse_course_flag(name: str) -> CourseFlag:
    lower = name.lower()
    if lower.startswith("ap ") or lower.startswith("g ap "):
        return CourseFlag.AP
    if "honors" in lower or "hon " in lower:
        return CourseFlag.HONORS
    if "college" in lower:
        return CourseFlag.COLLEGE
    return CourseFlag.REGULAR


def _detect_gender_restriction(title: str) -> Optional[Gender]:
    if title.startswith("G "):
        return Gender.GIRLS
    return None


def _detect_department(title: str) -> str:
    lower = title.lower().replace("g ", "", 1) if title.startswith("G ") else title.lower()

    dept_map = [
        (["sacred scrip", "sacrament", "jesus", "world relig", "advanced liturgy", "redeemer"], "Religion"),
        (["english", "american lit", "world lit", "grammar", "genre", "creative writ", "composition", "journalism"], "English"),
        (["algebra", "geometry", "calculus", "pre-calc", "pre calc", "statistics", "stats", "math analysis"], "Math"),
        (["biology", "chemistry", "physics", "anatomy", "genetics", "environmental", "forensic sci", "conceptual phys", "college bio", "nursing", "engineering"], "Science"),
        (["history", "geography", "econ", "government", "civil war", "current events", "moments in", "u.s. hist", "world hist", "women in lead", "polish hist"], "History"),
        (["spanish", "french", "polish"], "Language"),
        (["pe", "health"], "PE"),
        (["art", "multimedia", "studio art"], "Arts"),
        (["band", "choir", "music"], "Music"),
        (["computer science", "java"], "Computer Science"),
        (["business", "accounting", "deca", "building wealth", "act prep"], "Business"),
        (["psychology", "psych"], "Psychology"),
        (["seminar"], "Seminar"),
        (["forensics/debate", "speech"], "Elective"),
    ]

    for keywords, dept in dept_map:
        if any(kw in lower for kw in keywords):
            return dept
    return "Elective"


def _detect_grade_level(title: str, requesting_grades: set[int]) -> int:
    lower = title.lower()
    if " 9" in lower or "english 9" in lower or "scrip" in lower:
        return 9
    if " 10" in lower or "english 10" in lower or "sacrament" in lower:
        return 10
    if "american lit" in lower or "redeemer" in lower or "u.s. hist" in lower:
        return 11
    if "world lit" in lower or "world relig" in lower or "econ" in lower or "liturgy 12" in lower:
        return 12
    if "seminar 100" in lower:
        return 9
    if "seminar 200" in lower:
        return 10
    if "seminar 300" in lower:
        return 11
    if "seminar 400" in lower:
        return 12
    if requesting_grades:
        return min(requesting_grades)
    return 0


# ── CSV PARSER (Student_Requests.csv) ───────────────────────────────────

def load_student_requests_csv(filepath: str | Path, rotation_girls: str = "",
                               rotation_boys: str = "") -> ScheduleData:
    filepath = Path(filepath)
    data = ScheduleData()

    with open(filepath, encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        rows = list(reader)

    student_map: dict[str, dict] = {}
    course_counter: Counter = Counter()
    course_grades: dict[str, set[int]] = defaultdict(set)
    course_genders: dict[str, set[str]] = defaultdict(set)

    for row in rows:
        name = row.get("Student summary", "").strip()
        grade_str = row.get("Student grade level", "").strip()
        title = row.get("Title", "").strip()
        gender_str = row.get("Gender", "").strip()

        if not name or not title:
            continue

        grade = _parse_grade(grade_str)
        gender = _parse_gender(gender_str)
        if not gender:
            gender_from_grade = grade_str.lower()
            if "female" in gender_from_grade:
                gender = Gender.GIRLS
            elif "male" in gender_from_grade:
                gender = Gender.BOYS

        if name not in student_map:
            sid = str(10001 + len(student_map))
            parts = name.split(None, 1)
            first = parts[0] if parts else name
            last = parts[1] if len(parts) > 1 else ""
            student_map[name] = {
                "id": sid, "first": first, "last": last,
                "grade": grade, "gender": gender,
                "courses": [],
            }

        sdata = student_map[name]
        if grade > 0:
            sdata["grade"] = grade
        if gender:
            sdata["gender"] = gender
        sdata["courses"].append(title)

        course_counter[title] += 1
        if grade > 0:
            course_grades[title].add(grade)
        if gender:
            course_genders[title].add(gender.value)

    for name, sdata in student_map.items():
        if not sdata["gender"]:
            continue
        data.students[sdata["id"]] = Student(
            student_id=sdata["id"],
            last_name=sdata["last"],
            first_name=sdata["first"],
            grade=sdata["grade"],
            gender=sdata["gender"],
            course_requests=sdata["courses"],
        )

    rot_girls = rotation_girls or "G Grammar and Genre Studies 9"
    rot_boys = rotation_boys or "Grammar and Genre Studies"

    for title, count in course_counter.items():
        gender_restriction = _detect_gender_restriction(title)
        grades = course_grades.get(title, set())
        grade_level = _detect_grade_level(title, grades)
        dept = _detect_department(title)
        flag = _parse_course_flag(title)

        is_rotation = (title == rot_girls or title == rot_boys)
        if is_rotation:
            flag = CourseFlag.ROTATION

        is_seminar = "seminar" in title.lower()

        capacity = 24
        if dept == "PE":
            capacity = 30
        elif dept == "Arts":
            capacity = 20
        elif is_seminar:
            capacity = 30
        elif is_rotation:
            capacity = 55

        min_sections = max(1, math.ceil(count / capacity))

        data.courses[title] = Course(
            code=title,
            name=title,
            department=dept,
            grade_level=grade_level,
            gender_restriction=gender_restriction,
            min_sections=min_sections,
            standard_capacity=capacity,
            flag=flag,
            enrollment_count=count,
        )

    _add_default_constraints(data)
    _auto_generate_teachers(data)

    print(f"  Parsed {len(data.students)} students, {len(data.courses)} courses, {len(data.teachers)} teachers (auto)")
    print(f"  Rotation (girls): '{rot_girls}' — {course_counter.get(rot_girls, 0)} students")
    print(f"  Rotation (boys): '{rot_boys}' — {course_counter.get(rot_boys, 0)} students")

    return data


def _auto_generate_teachers(data: ScheduleData):
    """Generate placeholder teachers when no teacher roster is provided."""
    if data.teachers:
        return

    dept_sections: dict[str, int] = defaultdict(int)
    dept_courses: dict[str, list[str]] = defaultdict(list)

    for course in data.courses.values():
        if course.enrollment_count == 0:
            continue
        sections_needed = course.required_sections()
        dept = course.department
        dept_sections[dept] += sections_needed
        dept_courses[dept].append(course.code)

    teacher_id = 1
    last_names = [
        "Smith", "Johnson", "Williams", "Brown", "Jones", "Davis", "Miller",
        "Wilson", "Moore", "Taylor", "Anderson", "Thomas", "Jackson", "White",
        "Harris", "Martin", "Thompson", "Garcia", "Martinez", "Robinson",
        "Clark", "Rodriguez", "Lewis", "Lee", "Walker", "Hall", "Allen",
        "Young", "King", "Wright", "Lopez", "Hill", "Scott", "Green",
        "Adams", "Baker", "Nelson", "Carter", "Mitchell", "Perez",
        "Roberts", "Turner", "Phillips", "Campbell", "Parker", "Evans",
        "Edwards", "Collins", "Stewart", "Reed", "Cooper", "Morgan",
        "Kelly", "Howard", "Ward", "Cox", "Diaz", "Richardson",
        "Wood", "Watson", "Brooks", "Bennett", "Gray", "Price",
    ]
    first_names = [
        "James", "Mary", "Robert", "Patricia", "John", "Jennifer", "Michael",
        "Linda", "David", "Elizabeth", "William", "Barbara", "Richard", "Susan",
        "Joseph", "Jessica", "Thomas", "Sarah", "Christopher", "Karen",
        "Charles", "Lisa", "Daniel", "Nancy", "Matthew", "Betty", "Anthony",
        "Margaret", "Mark", "Sandra", "Steven", "Ashley", "Paul", "Dorothy",
        "Andrew", "Kimberly", "Joshua", "Emily", "Kenneth", "Donna",
    ]

    for dept, total_sections in sorted(dept_sections.items()):
        teachers_needed = max(1, math.ceil(total_sections / 6))
        courses_in_dept = dept_courses[dept]
        specialties = courses_in_dept[:6]

        is_rotation_dept = dept.lower() in ("rotation",)
        is_seminar_dept = dept.lower() == "seminar"

        for i in range(teachers_needed):
            last = last_names[(teacher_id - 1) % len(last_names)]
            first = first_names[(teacher_id - 1) % len(first_names)]
            name = f"{last}, {first}"

            if name in data.teachers:
                name = f"{last}, {first} {teacher_id}"

            certs = []
            for spec in specialties:
                if "ap " in spec.lower() or spec.lower().startswith("ap "):
                    certs.append(f"AP {spec}")
                if "honor" in spec.lower():
                    certs.append(f"Honors {spec}")

            max_tp = 6
            prep = 1
            if is_rotation_dept:
                max_tp = 5
                prep = 2
            if is_seminar_dept:
                max_tp = 6

            data.teachers[name] = Teacher(
                name=name,
                department=dept,
                subject_specialties=specialties,
                certifications=certs,
                max_teaching_periods=max_tp,
                required_prep_periods=prep,
                is_rotation_teacher=is_rotation_dept,
                notes="Auto-generated" + (" Both pools" if is_rotation_dept else ""),
            )
            teacher_id += 1


def _add_default_constraints(data: ScheduleData):
    data.constraints = [
        Constraint("Rotation Period", "Girls 9th Grade", "Section A = P1, Section B = P2", "Built-in", "LOCKED"),
        Constraint("Rotation Period", "Boys 10th Grade", "Section A = P3, Section B = P5B", "Built-in", "LOCKED"),
        Constraint("Lunch Time", "All 9-10 students", "P5A", "Built-in", "LOCKED"),
        Constraint("Lunch Time", "All 11-12 students", "P5B", "Built-in", "LOCKED"),
        Constraint("SEM Period", "All", "P4 — Seminar by grade", "Built-in", "LOCKED"),
    ]


# ── XLSX PARSER (INPUT_TEMPLATES.xlsx) ──────────────────────────────────

def _parse_periods(val: str) -> list[Period]:
    if not val or val.lower() == "all":
        return list(ALL_PERIODS)
    period_map = {p.value.upper(): p for p in Period}
    result = []
    for part in val.replace(",", " ").split():
        part = part.strip().upper()
        if part in period_map:
            result.append(period_map[part])
    return result or list(ALL_PERIODS)


def _split_list(val: str) -> list[str]:
    if not val:
        return []
    return [x.strip() for x in val.replace(";", ",").split(",") if x.strip()]


def _find_header_row(ws, expected_headers: list[str]) -> int:
    for row_idx in range(1, min(ws.max_row + 1, 10)):
        values = [_str(ws.cell(row=row_idx, column=c).value).lower()
                  for c in range(1, ws.max_column + 1)]
        matches = sum(1 for h in expected_headers if any(h in v for v in values))
        if matches >= len(expected_headers) // 2:
            return row_idx
    return 0


def _col_map(ws, header_row: int) -> dict[str, int]:
    mapping = {}
    for c in range(1, ws.max_column + 1):
        val = _str(ws.cell(row=header_row, column=c).value).lower()
        if val:
            mapping[val] = c
    return mapping


def _get(ws, row: int, col_map: dict, *keys: str) -> str:
    for key in keys:
        for col_name, col_idx in col_map.items():
            if key in col_name:
                return _str(ws.cell(row=row, column=col_idx).value)
    return ""


def parse_students_xlsx(ws) -> dict[str, Student]:
    header_row = _find_header_row(ws, ["student id", "last name", "grade", "gender"])
    if not header_row:
        raise ValueError("Cannot find header row in Student Selections sheet")

    cmap = _col_map(ws, header_row)
    students = {}

    for row in range(header_row + 1, ws.max_row + 1):
        sid = _get(ws, row, cmap, "student id", "id")
        if not sid:
            continue

        last_name = _get(ws, row, cmap, "last name", "last")
        first_name = _get(ws, row, cmap, "first name", "first")
        grade = _parse_grade(_get(ws, row, cmap, "grade"))
        gender = _parse_gender(_get(ws, row, cmap, "gender", "sex"))

        if not gender or grade == 0:
            continue

        courses = []
        for col_idx in range(1, ws.max_column + 1):
            col_header = _str(ws.cell(row=header_row, column=col_idx).value).lower()
            if "course" in col_header:
                course_val = _str(ws.cell(row=row, column=col_idx).value)
                if course_val:
                    courses.append(course_val)

        students[sid] = Student(
            student_id=sid, last_name=last_name, first_name=first_name,
            grade=grade, gender=gender, course_requests=courses,
        )

    return students


def parse_teachers_xlsx(ws) -> dict[str, Teacher]:
    header_row = _find_header_row(ws, ["teacher name", "department"])
    if not header_row:
        raise ValueError("Cannot find header row in Teacher Roster sheet")

    cmap = _col_map(ws, header_row)
    teachers = {}

    for row in range(header_row + 1, ws.max_row + 1):
        name = _get(ws, row, cmap, "teacher name", "name")
        if not name:
            continue

        dept = _get(ws, row, cmap, "department", "dept")
        specialties = _split_list(_get(ws, row, cmap, "subject", "specialt"))
        certs = _split_list(_get(ws, row, cmap, "certification", "cert"))
        max_tp = _int(_get(ws, row, cmap, "max teach", "max period"), 6)
        prep = _int(_get(ws, row, cmap, "prep", "required prep"), 1)
        constraints = _get(ws, row, cmap, "constraint", "special")
        notes = _get(ws, row, cmap, "note")
        is_rotation = "rotation" in constraints.lower() or "rotation" in dept.lower()

        teachers[name] = Teacher(
            name=name, department=dept, subject_specialties=specialties,
            certifications=certs, max_teaching_periods=max_tp,
            required_prep_periods=prep, special_constraints=constraints,
            notes=notes, is_rotation_teacher=is_rotation,
        )

    return teachers


def parse_rooms_xlsx(ws) -> dict[str, Room]:
    header_row = _find_header_row(ws, ["room number", "room type", "capacity"])
    if not header_row:
        raise ValueError("Cannot find header row in Room Inventory sheet")

    cmap = _col_map(ws, header_row)
    rooms = {}

    for row in range(header_row + 1, ws.max_row + 1):
        number = _get(ws, row, cmap, "room number", "room")
        if not number:
            continue
        rtype = _get(ws, row, cmap, "room type", "type")
        cap = _int(_get(ws, row, cmap, "capacity", "cap"), 30)
        features = _get(ws, row, cmap, "feature", "special")
        avail = _parse_periods(_get(ws, row, cmap, "available", "period"))
        notes = _get(ws, row, cmap, "note")

        rooms[number] = Room(
            number=number, room_type=rtype, capacity=cap,
            special_features=features, available_periods=avail, notes=notes,
        )

    return rooms


def parse_constraints_xlsx(ws) -> list[Constraint]:
    header_row = _find_header_row(ws, ["constraint", "subject", "detail"])
    if not header_row:
        raise ValueError("Cannot find header row in Constraints sheet")

    cmap = _col_map(ws, header_row)
    constraints = []

    for row in range(header_row + 1, ws.max_row + 1):
        ctype = _get(ws, row, cmap, "constraint type", "type")
        if not ctype:
            continue
        subject = _get(ws, row, cmap, "subject", "teacher")
        detail = _get(ws, row, cmap, "detail")
        approval = _get(ws, row, cmap, "approval")
        status = _get(ws, row, cmap, "status")
        constraints.append(Constraint(
            constraint_type=ctype, subject_teacher=subject,
            detail=detail, approval_required_from=approval, status=status,
        ))

    return constraints


def load_input_file(filepath: str | Path) -> ScheduleData:
    filepath = Path(filepath)
    if not filepath.exists():
        raise FileNotFoundError(f"Input file not found: {filepath}")

    if filepath.suffix.lower() == ".csv":
        return load_student_requests_csv(filepath)

    # Check if this is a Teacher Roster v16 file
    if _is_teacher_roster_v16(filepath):
        data = ScheduleData()
        parse_teacher_roster_v16(filepath, data)
        return data

    wb = openpyxl.load_workbook(filepath, data_only=True)
    data = ScheduleData()

    sheet_map = {}
    for name in wb.sheetnames:
        lower = name.lower()
        if "student" in lower or "selection" in lower:
            sheet_map["students"] = name
        elif "course" in lower or "catalog" in lower:
            sheet_map["courses"] = name
        elif "teacher" in lower or "roster" in lower:
            sheet_map["teachers"] = name
        elif "room" in lower or "inventory" in lower:
            sheet_map["rooms"] = name
        elif "constraint" in lower:
            sheet_map["constraints"] = name

    if "students" in sheet_map:
        data.students = parse_students_xlsx(wb[sheet_map["students"]])
    if "teachers" in sheet_map:
        data.teachers = parse_teachers_xlsx(wb[sheet_map["teachers"]])
    if "rooms" in sheet_map:
        data.rooms = parse_rooms_xlsx(wb[sheet_map["rooms"]])
    if "constraints" in sheet_map:
        data.constraints = parse_constraints_xlsx(wb[sheet_map["constraints"]])

    if "courses" in sheet_map:
        ws = wb[sheet_map["courses"]]
        header_row = _find_header_row(ws, ["course code", "course name", "department"])
        if header_row:
            cmap = _col_map(ws, header_row)
            for row in range(header_row + 1, ws.max_row + 1):
                code = _get(ws, row, cmap, "course code", "code")
                if not code:
                    continue
                name = _get(ws, row, cmap, "course name", "name")
                dept = _get(ws, row, cmap, "department", "dept")
                grade = _parse_grade(_get(ws, row, cmap, "grade"))
                gender_str = _get(ws, row, cmap, "gender", "restriction")
                gender = _parse_gender(gender_str) if gender_str else None
                min_sec = _int(_get(ws, row, cmap, "min section", "sections"), 1)
                cap = _int(_get(ws, row, cmap, "capacity", "cap"), 24)
                flag_str = _get(ws, row, cmap, "honor", "ap", "flag")
                flag = _parse_course_flag(flag_str) if flag_str else CourseFlag.REGULAR

                data.courses[code] = Course(
                    code=code, name=name, department=dept, grade_level=grade,
                    gender_restriction=gender, min_sections=min_sec,
                    standard_capacity=cap, flag=flag,
                )

    _compute_enrollment_counts(data)
    return data


def _compute_enrollment_counts(data: ScheduleData):
    for course in data.courses.values():
        course.enrollment_count = 0
    for student in data.students.values():
        for req in student.course_requests:
            if req in data.courses:
                data.courses[req].enrollment_count += 1


# ── TEACHER ROSTER v16 PARSER ─────────────────────────────────────────

def _parse_period(val: str) -> Optional[Period]:
    """Parse a single period string like 'P5A' into a Period enum."""
    val = val.strip().upper()
    try:
        return Period(val)
    except ValueError:
        return None


def _parse_prep_list(val: str) -> list[Period]:
    """Parse a comma-separated list of PREP periods."""
    if not val or val.strip() == "—" or val.strip() == "-":
        return []
    result = []
    for part in val.replace(",", " ").split():
        p = _parse_period(part)
        if p:
            result.append(p)
    return result


def _resolve_teacher_name(name: str, teachers: dict[str, Teacher]) -> str:
    """Resolve a teacher name to its canonical form in the teachers dict.

    Handles mismatches like 'Carly (PT Hire)' vs 'Carly (NEW PT HIRE)' by
    matching on last-name prefix when exact match fails.
    """
    if name in teachers:
        return name
    # Try matching by first token (last name or unique identifier)
    name_lower = name.lower()
    first_token = name.split(",")[0].split("(")[0].strip().lower()
    for known in teachers:
        known_first = known.split(",")[0].split("(")[0].strip().lower()
        if first_token and known_first and first_token == known_first:
            # Ambiguity check: only match if there's exactly one candidate
            candidates = [k for k in teachers
                          if k.split(",")[0].split("(")[0].strip().lower() == first_token]
            if len(candidates) == 1:
                return candidates[0]
    return name  # Return original if no match found


def _split_qualified_teachers(raw: str, known_names: set[str]) -> list[str]:
    """Split a comma-separated string of qualified teachers into individual names.

    Teacher names use "Last, First" format (e.g. "Abel, Mike") and multiple
    teachers are also comma-separated. So "Graham, Ethan, Lengel, Mark"
    must parse as ["Graham, Ethan", "Lengel, Mark"].

    Strategy: use the known teacher names from the master roster to greedily
    match names from the raw string. Fall back to pairing tokens for unknown names.
    """
    if not raw:
        return []

    # Try to greedily match known names
    remaining = raw.strip()
    result: list[str] = []

    # Sort known names longest-first so we match the most specific name first
    sorted_names = sorted(known_names, key=len, reverse=True)

    while remaining:
        remaining = remaining.strip().lstrip(",").strip()
        if not remaining:
            break

        matched = False
        for name in sorted_names:
            if remaining.startswith(name):
                rest = remaining[len(name):]
                # Must be followed by end-of-string, comma, or whitespace
                if not rest or rest[0] in (",", " "):
                    result.append(name)
                    remaining = rest.lstrip(",").strip()
                    matched = True
                    break

        if not matched:
            # Unknown name — take up to the next comma-separated pair
            # Assume "Last, First" format: take two comma-separated tokens
            parts = remaining.split(",", 2)
            if len(parts) >= 2:
                candidate = (parts[0].strip() + ", " + parts[1].strip())
                result.append(candidate)
                remaining = ",".join(parts[2:]) if len(parts) > 2 else ""
            else:
                # Single token remaining
                if parts[0].strip():
                    result.append(parts[0].strip())
                remaining = ""

    return result


def _build_fuzzy_course_map(roster_courses: list[str],
                            student_courses: list[str]) -> dict[str, str]:
    """Build a mapping from student CSV course names to roster short names.

    The roster uses abbreviated names like "G Hon Bio", "AP Amer Lit B"
    while the student CSV uses full names like "G Honors Biology",
    "AP American Literature".  We build a bidirectional fuzzy lookup.

    Returns dict mapping roster_course -> student_course (best match).
    """
    import re

    roman_map = {"i": "1", "ii": "2", "iii": "3", "iv": "4"}

    def _normalize(name: str) -> str:
        """Normalize a course name for comparison."""
        s = name.lower().strip()
        s = s.replace("(combined)", "").replace("(r)", "")
        s = s.replace("/", " ").replace("-", " ").replace("&", "and")
        s = s.replace("u.s.", "us")
        for roman, digit in roman_map.items():
            s = re.sub(rf'\b{roman}\b', digit, s)
        return re.sub(r'\s+', ' ', s).strip()

    def _tokenize(name: str) -> set[str]:
        """Get meaningful tokens from a course name."""
        s = _normalize(name)
        noise = {"and", "the", "of", "in", "to", "for"}
        return {w for w in s.split() if (len(w) > 1 or w.isdigit()) and w not in noise}

    abbrev_map = {
        "amer": "american",
        "lit": "literature",
        "bio": "biology",
        "chem": "chemistry",
        "phys": "physics",
        "conc": "conceptual",
        "scrip": "scriptures",
        "calc": "calculus",
        "alg": "algebra",
        "geo": "geometry",
        "hon": "honors",
        "eng": "english",
        "hist": "history",
        "relig": "religions",
        "psyc": "psychology",
        "comp": "composition",
        "env": "environmental",
        "sci": "science",
        "wrld": "world",
        "econ": "economics",
        "acct": "accounting",
        "account": "accounting",
        "stats": "statistics",
        "prep": "preparation",
        "mm": "multimedia",
        "rot": "rotation",
        "redeem": "redeemer",
        "gov": "government",
        "pe": "physical education",
        "adv": "advanced",
        "pre": "pre",
    }

    def _expand(tokens: set[str]) -> set[str]:
        expanded = set(tokens)
        for tok in tokens:
            if tok in abbrev_map:
                expanded.add(abbrev_map[tok])
        return expanded

    mapping: dict[str, str] = {}
    used_student: set[str] = set()

    # First pass: exact matches
    student_lower_map = {s.lower(): s for s in student_courses}
    for rc in roster_courses:
        if rc.lower() in student_lower_map:
            mapping[rc] = student_lower_map[rc.lower()]
            used_student.add(rc.lower())

    # Second pass: fuzzy matching for unmatched roster courses
    remaining_roster = [rc for rc in roster_courses if rc not in mapping]
    remaining_student = [s for s in student_courses if s.lower() not in used_student]

    for rc_orig in remaining_roster:
        rc = rc_orig
        rc_tokens = _tokenize(rc)
        rc_expanded = _expand(rc_tokens)

        best_score = 0
        best_match = None

        for sc in remaining_student:
            sc_tokens = _tokenize(sc)
            sc_expanded = _expand(sc_tokens)

            # Check gender prefix/suffix match
            rc_has_g = rc.startswith("G ") or (rc.endswith(" G") and not rc.endswith("OG"))
            sc_has_g = sc.startswith("G ") or (sc.endswith(" G") and not sc.endswith("OG"))
            rc_combined = "(combined)" in rc_orig.lower() or "(Combined)" in rc_orig
            rc_has_b = rc.endswith(" B") and not rc.endswith("AB")
            sc_has_b = sc.endswith(" B") and not sc.endswith("AB")

            if rc_combined:
                pass
            elif rc_has_g != sc_has_g:
                continue
            elif rc_has_b and sc_has_g:
                continue
            elif sc_has_b and rc_has_g:
                continue

            # Token overlap with expansion (exclude gender/boys suffix tokens)
            rc_match_tokens = rc_expanded - {"b"} if rc_has_b else rc_expanded
            sc_match_tokens = sc_expanded - {"b"} if sc_has_b else sc_expanded
            if rc.endswith(" G"):
                rc_match_tokens = rc_match_tokens - {"g"}
            if sc.endswith(" G"):
                sc_match_tokens = sc_match_tokens - {"g"}
            overlap = len(rc_match_tokens & sc_match_tokens)
            union = len(rc_match_tokens | sc_match_tokens)
            if union == 0:
                continue

            score = overlap / union

            # Bonus for matching B suffix
            if rc_has_b == sc_has_b:
                score += 0.05

            # Bonus for AP match
            rc_is_ap = "ap" in rc_tokens
            sc_is_ap = "ap" in sc_tokens
            if rc_is_ap == sc_is_ap:
                score += 0.1
            elif rc_is_ap != sc_is_ap:
                score -= 0.5

            if score > best_score:
                best_score = score
                best_match = sc

        if best_match and best_score > 0.25:
            mapping[rc_orig] = best_match
            if not rc_combined:
                used_student.add(best_match.lower())

    return mapping


def parse_teacher_roster_v16(filepath: str | Path, data: ScheduleData) -> None:
    """Parse the OLSM Teacher Roster v16 Excel file.

    Reads four sheets:
    1. "1. Master Roster" -> data.teachers
    2. "3. Rotation Pools" -> data.girls_rotation_teachers, data.boys_rotation_teachers
    3. "5. Constraints" -> overload/lunch shift adjustments on teachers
    4. "6. Subject Specialties" -> data._teacher_course_map
    """
    filepath = Path(filepath)
    wb = openpyxl.load_workbook(filepath, data_only=True)

    # ── Sheet 1: Master Roster ──────────────────────────────────────
    ws1 = wb["1. Master Roster"]
    # Header is at row 4, data starts at row 5
    header_row = 4
    cmap1 = _col_map(ws1, header_row)

    for row in range(5, ws1.max_row + 1):
        name = _get(ws1, row, cmap1, "last name", "first")
        if not name:
            continue

        dept = _get(ws1, row, cmap1, "department", "dept")
        max_tp = _int(_get(ws1, row, cmap1, "total teaching"), 6)

        # Subject specialties from "Subjects Taught (Unique)" column
        subjects_raw = _get(ws1, row, cmap1, "subjects taught")
        specialties = [s.strip() for s in subjects_raw.replace(";", ",").split(",")
                       if s.strip() and s.strip() != "—" and s.strip() != "— CUT —"]

        # Rotation pool
        rotation_raw = _get(ws1, row, cmap1, "rotation")
        is_rotation = bool(rotation_raw
                           and rotation_raw.strip() != "—"
                           and rotation_raw.strip() != "-"
                           and rotation_raw.strip())
        rotation_pools = []
        if is_rotation:
            if "girls" in rotation_raw.lower():
                rotation_pools.append("Girls")
            if "boys" in rotation_raw.lower():
                rotation_pools.append("Boys")

        # PREP periods
        prep_raw = _get(ws1, row, cmap1, "prep period")
        prep_periods = _parse_prep_list(prep_raw)
        required_prep = len(prep_periods) if prep_periods else 1

        # Lunch
        lunch_raw = _get(ws1, row, cmap1, "lunch")
        lunch_period = None
        if lunch_raw and lunch_raw.strip() != "—" and lunch_raw.strip() != "-":
            lunch_period = _parse_period(lunch_raw.split("(")[0].strip())

        # Overload
        overload_raw = _get(ws1, row, cmap1, "overload")
        overload = overload_raw.strip().upper() == "YES" if overload_raw else False

        # Notes
        notes = _get(ws1, row, cmap1, "notes", "constraint")

        data.teachers[name] = Teacher(
            name=name,
            department=dept or "",
            subject_specialties=specialties,
            certifications=[],
            max_teaching_periods=max_tp,
            required_prep_periods=required_prep,
            special_constraints="",
            notes=notes,
            is_rotation_teacher=is_rotation,
            rotation_pools=rotation_pools,
            lunch_period=lunch_period,
            overload_approved=overload,
        )

    # ── Sheet 3: Rotation Pools ─────────────────────────────────────
    ws3 = wb["3. Rotation Pools"]
    for row in range(5, ws3.max_row + 1):
        pool_val = _str(ws3.cell(row=row, column=1).value)
        teacher_name = _str(ws3.cell(row=row, column=3).value)
        if not teacher_name:
            continue

        # Resolve name to master roster key (handles mismatches like
        # "Carly (PT Hire)" in rotation pool vs "Carly (NEW PT HIRE)" in roster)
        resolved_name = _resolve_teacher_name(teacher_name, data.teachers)

        if "girls" in pool_val.lower() or "p1" in pool_val.lower():
            if resolved_name not in data.girls_rotation_teachers:
                data.girls_rotation_teachers.append(resolved_name)
        if "boys" in pool_val.lower() or "p3" in pool_val.lower():
            if resolved_name not in data.boys_rotation_teachers:
                data.boys_rotation_teachers.append(resolved_name)

        # Also update the teacher record if it exists
        if resolved_name in data.teachers:
            t = data.teachers[resolved_name]
            t.is_rotation_teacher = True

    # ── Sheet 5: Constraints ────────────────────────────────────────
    ws5 = wb["5. Constraints"]
    for row in range(4, ws5.max_row + 1):
        teacher_name = _str(ws5.cell(row=row, column=1).value)
        constraint_type = _str(ws5.cell(row=row, column=2).value).upper()
        detail = _str(ws5.cell(row=row, column=3).value)

        if not teacher_name or not constraint_type:
            continue

        if teacher_name not in data.teachers:
            continue

        t = data.teachers[teacher_name]

        if constraint_type == "OVERLOAD":
            t.overload_approved = True
            # Parse the max periods from detail like "8 teaching periods (standard is 6)"
            import re
            m = re.match(r'(\d+)\s+teaching', detail)
            if m:
                t.max_teaching_periods = int(m.group(1))

        elif constraint_type == "LUNCH SHIFT":
            # Parse the shifted lunch period from detail
            # "Lunch at P5A instead of standard P5B"
            import re
            m = re.search(r'Lunch at (P\w+)', detail)
            if m:
                p = _parse_period(m.group(1))
                if p:
                    t.lunch_period = p

    # ── Sheet 6: Subject Specialties ────────────────────────────────
    ws6 = wb["6. Subject Specialties"]
    teacher_course_map: dict[str, list[str]] = {}

    # Build set of known teacher names from master roster for smart splitting
    known_names = set(data.teachers.keys())

    for row in range(5, ws6.max_row + 1):
        course_name = _str(ws6.cell(row=row, column=1).value)
        qualified_raw = _str(ws6.cell(row=row, column=4).value)

        if not course_name or course_name == "— CUT —":
            continue

        teachers_list = _split_qualified_teachers(qualified_raw, known_names)
        teacher_course_map[course_name] = teachers_list

    # Store on the data object
    data._teacher_course_map = teacher_course_map  # type: ignore[attr-defined]

    # Build the fuzzy mapping from roster course names to student CSV course names
    # This will be populated later when student data is available
    data._roster_course_names = list(teacher_course_map.keys())  # type: ignore[attr-defined]

    # Pre-build a reverse map: teacher_name -> list of roster course names
    teacher_to_courses: dict[str, list[str]] = defaultdict(list)
    for course, teachers in teacher_course_map.items():
        for t_name in teachers:
            teacher_to_courses[t_name].append(course)
    data._teacher_to_roster_courses = dict(teacher_to_courses)  # type: ignore[attr-defined]

    print(f"  Parsed {len(data.teachers)} teachers from Teacher Roster v16")
    print(f"  Girls rotation pool: {data.girls_rotation_teachers}")
    print(f"  Boys rotation pool: {data.boys_rotation_teachers}")
    print(f"  Subject specialties: {len(teacher_course_map)} courses mapped")


def _is_teacher_roster_v16(filepath: str | Path) -> bool:
    """Check if a file is a Teacher Roster v16 file."""
    try:
        wb = openpyxl.load_workbook(filepath, read_only=True)
        result = "1. Master Roster" in wb.sheetnames
        wb.close()
        return result
    except Exception:
        return False


def build_fuzzy_course_mapping(data: ScheduleData) -> dict[str, str]:
    """Build fuzzy mapping between roster course names and student course names.

    Call this AFTER both teacher roster and student data are loaded.
    Returns dict mapping roster_course_name -> student_course_name.
    Also stores the mapping as data._course_name_map.
    """
    roster_courses = getattr(data, "_roster_course_names", [])
    if not roster_courses:
        return {}

    import re

    student_courses = list(data.courses.keys())
    mapping = _build_fuzzy_course_map(roster_courses, student_courses)
    data._course_name_map = mapping  # type: ignore[attr-defined]

    # Also build reverse: student_course -> roster_course
    reverse = {v: k for k, v in mapping.items()}

    tcm = getattr(data, "_teacher_course_map", {})

    # For "(combined)" roster courses, also map the G-prefix student equivalent
    for rc in roster_courses:
        if "(combined)" not in rc.lower() and "(Combined)" not in rc:
            continue
        if rc not in mapping:
            continue
        sc = mapping[rc]
        teachers = tcm.get(rc, [])
        if not teachers:
            continue
        # If we mapped "French 3 (Combined)" -> "French III", also map "G French III"
        g_variant = f"G {sc}" if not sc.startswith("G ") else None
        non_g_variant = sc[2:] if sc.startswith("G ") else None
        for variant in [g_variant, non_g_variant]:
            if variant and variant in data.courses and variant not in reverse:
                reverse[variant] = rc

    # Handle range roster entries like "Band 1-4"
    roman_rev = {"1": "I", "2": "II", "3": "III", "4": "IV"}
    for rc in roster_courses:
        m = re.match(r'^(.+?)\s*(\d)\s*-\s*(\d)\s*$', rc)
        if not m:
            continue
        base, start, end = m.group(1).strip(), int(m.group(2)), int(m.group(3))
        teachers = tcm.get(rc, [])
        if not teachers:
            continue
        for num in range(start, end + 1):
            for sc in student_courses:
                if sc in reverse:
                    continue
                sc_lower = sc.lower()
                base_lower = base.lower()
                numeral = roman_rev.get(str(num), str(num))
                if base_lower in sc_lower and (numeral in sc or str(num) in sc):
                    reverse[sc] = rc
                    break

    # Department synonym groups
    dept_synonyms = {
        "language": {"language", "french", "spanish", "world lang"},
        "french": {"language", "french"},
        "spanish": {"language", "spanish"},
        "science": {"science", "biology", "chemistry", "physics", "engineering"},
        "history": {"history", "social studies", "government", "economics"},
        "pe": {"pe", "health", "physical education"},
        "arts": {"arts", "art", "music", "fine arts"},
        "computer science": {"computer", "computer science", "cs", "technology"},
        "math": {"math", "mathematics"},
        "english": {"english", "language arts", "writing"},
        "business": {"business", "accounting"},
        "music": {"music", "arts", "fine arts"},
    }

    def _depts_match(d1: str, d2: str) -> bool:
        if d1 == d2 or d1 in d2 or d2 in d1:
            return True
        syns1 = dept_synonyms.get(d1, {d1})
        syns2 = dept_synonyms.get(d2, {d2})
        return bool(syns1 & syns2)

    # Department-based fallback for unmapped courses with >0 enrollment
    for sc in student_courses:
        if sc in reverse:
            continue
        course = data.courses.get(sc)
        if not course or course.enrollment_count == 0:
            continue
        if "seminar" in sc.lower() or course.is_rotation:
            continue
        course_dept = course.department.lower().strip()
        if not course_dept:
            continue
        for t in data.teachers.values():
            t_dept = t.department.lower().split("/")[0].strip()
            if not t_dept:
                continue
            if _depts_match(course_dept, t_dept):
                if sc not in reverse:
                    reverse[sc] = f"__dept__{sc}"
                    tcm[f"__dept__{sc}"] = []
                if t.name not in tcm[f"__dept__{sc}"]:
                    tcm[f"__dept__{sc}"].append(t.name)

    data._student_to_roster_course = reverse  # type: ignore[attr-defined]

    return mapping
