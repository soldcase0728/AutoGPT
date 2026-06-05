"""Generate realistic sample data matching OLSM's ~847 student body."""

from __future__ import annotations

import random
import openpyxl
from pathlib import Path


GIRL_FIRST_NAMES = [
    "Emma", "Olivia", "Ava", "Isabella", "Sophia", "Mia", "Charlotte", "Amelia",
    "Harper", "Evelyn", "Abigail", "Emily", "Elizabeth", "Sofia", "Avery",
    "Ella", "Scarlett", "Grace", "Victoria", "Riley", "Aria", "Lily", "Hannah",
    "Layla", "Zoey", "Nora", "Camila", "Penelope", "Luna", "Chloe",
    "Madison", "Eleanor", "Hazel", "Violet", "Aurora", "Savannah", "Audrey",
    "Brooklyn", "Bella", "Claire", "Skylar", "Lucy", "Paisley", "Anna",
    "Caroline", "Genesis", "Aaliyah", "Kennedy", "Kinsley", "Allison",
]

BOY_FIRST_NAMES = [
    "Liam", "Noah", "Oliver", "James", "Elijah", "William", "Henry", "Lucas",
    "Benjamin", "Theodore", "Jack", "Levi", "Alexander", "Mason", "Ethan",
    "Daniel", "Jacob", "Logan", "Jackson", "Sebastian", "Aiden", "Owen",
    "Samuel", "Ryan", "Nathan", "Carter", "Luke", "Jayden", "John", "David",
    "Andrew", "Isaiah", "Christopher", "Joshua", "Julian", "Michael", "Wyatt",
    "Matthew", "Dylan", "Caleb", "Hunter", "Connor", "Cameron", "Dominic",
    "Christian", "Thomas", "Aaron", "Charles", "Austin", "Marcus",
]

LAST_NAMES = [
    "Smith", "Johnson", "Williams", "Brown", "Jones", "Garcia", "Miller",
    "Davis", "Rodriguez", "Martinez", "Anderson", "Taylor", "Thomas", "Moore",
    "Jackson", "Martin", "Lee", "Perez", "Thompson", "White", "Harris",
    "Clark", "Lewis", "Robinson", "Walker", "Young", "King", "Wright",
    "Scott", "Green", "Baker", "Adams", "Nelson", "Hill", "Ramirez",
    "Campbell", "Mitchell", "Roberts", "Carter", "Phillips", "Evans",
    "Turner", "Torres", "Parker", "Collins", "Edwards", "Stewart", "Flores",
    "Morris", "Nguyen", "Murphy", "Rivera", "Cook", "Rogers", "Morgan",
    "Peterson", "Cooper", "Reed", "Bailey", "Bell", "Gomez", "Kelly",
    "Howard", "Ward", "Cox", "Diaz", "Richardson", "Wood", "Watson",
    "Brooks", "Bennett", "Gray", "James", "Reyes", "Cruz", "Hughes",
    "Price", "Myers", "Long", "Foster", "Sanders", "Ross", "Morales",
    "Powell", "Sullivan", "Russell", "Ortiz", "Jenkins", "Gutierrez",
    "Perry", "Butler", "Barnes", "Fisher", "Kowalski", "Nowak", "Mazur",
]

COURSES_9_GIRLS = {
    "religion": ["G Sacred Scrip"],
    "english": ["G Eng 9", "G Hon Eng 9"],
    "math": ["G Alg 1", "G Hon Geo"],
    "science": ["G Bio", "G Hon Bio"],
    "history": ["G World Hist"],
    "language": ["G Spanish 1", "G French 1"],
    "elective": ["Computer Science G", "G Polish", "G Art Apprec", "G Intro Music"],
}

COURSES_9_BOYS = {
    "religion": ["Sacred Scrip B"],
    "english": ["Eng 9 B", "Hon Eng 9 B"],
    "math": ["Alg 1 B", "Hon Geo B"],
    "science": ["Bio B", "Hon Bio B"],
    "history": ["World Hist B"],
    "language": ["Spanish 1 B", "French 1 B"],
    "elective": ["Computer Science B", "PE Health B", "Intro Music B"],
}

COURSES_10_GIRLS = {
    "religion": ["G Sacraments"],
    "english": ["G Eng 10", "G Hon Eng 10"],
    "math": ["G Geo", "G Alg 2", "G Hon Alg 2"],
    "science": ["G Chem", "G Hon Chem"],
    "history": ["G World Hist 2", "G AP World Hist"],
    "language": ["G Spanish 2", "G French 2"],
    "elective": ["G Art Apprec", "G Anatomy", "G Comp Sci 2", "G PE Health"],
}

COURSES_10_BOYS = {
    "religion": ["Sacraments B"],
    "english": ["Eng 10 B", "Hon Eng 10 B"],
    "math": ["Geo B", "Alg 2 B", "Hon Alg 2 B"],
    "science": ["Chem B", "Hon Chem B"],
    "history": ["World Hist B 2", "AP World Hist B"],
    "language": ["Spanish 2 B", "French 2 B"],
    "elective": ["Bio B Lab", "PE Health B 2", "Intro Business B", "Art B"],
}

COURSES_11 = {
    "religion": ["Jesus Redeem", "G Jesus Redeem"],
    "english": ["Amer Lit", "G Amer Lit", "AP Amer Lit", "G AP Amer Lit"],
    "math": ["Pre-Calc", "G Pre-Calc", "Hon Pre-Calc", "G Hon Pre-Calc"],
    "science": ["Physics", "G Physics", "AP Chem", "G AP Chem", "Anatomy", "G Anatomy"],
    "history": ["US Hist", "G US Hist", "AP US Hist", "G AP US Hist"],
    "language": ["Spanish 3", "G Spanish 3", "French 3", "G French 3"],
    "elective": ["Psych", "G Psych", "Art 2", "G Art 2", "Band", "Comp Sci AP"],
}

COURSES_12 = {
    "religion": ["World Relig", "G World Relig", "Adv Liturgy"],
    "english": ["World Lit", "G World Lit", "AP World Lit", "G AP World Lit"],
    "math": ["College Calc", "G College Calc", "Stats", "G Stats", "Math Analysis"],
    "science": ["College Bio", "G College Bio", "AP Physics", "Genetics", "G Genetics"],
    "history": ["Econ", "G Econ", "AP Econ"],
    "elective": ["Business", "G Business", "Psych 2", "G Psych 2", "AP Art", "Band 2", "Speech"],
}


def _pick_gender_course(options: list[str], gender: str) -> str:
    gender_filtered = [c for c in options if
                       (gender == "G" and ("G " in c or c.startswith("G "))) or
                       (gender == "B" and (" B" in c or c.endswith(" B")))]
    if gender_filtered:
        return random.choice(gender_filtered)
    return random.choice(options)


def generate_students(count: int = 847) -> list[dict]:
    students = []
    grade_dist = {9: 0.27, 10: 0.26, 11: 0.24, 12: 0.23}
    gender_dist = {"B": 0.52, "G": 0.48}

    student_id = 10001
    for grade, grade_pct in grade_dist.items():
        grade_count = round(count * grade_pct)
        for gender, gender_pct in gender_dist.items():
            gender_count = round(grade_count * gender_pct)
            names = GIRL_FIRST_NAMES if gender == "G" else BOY_FIRST_NAMES

            if grade == 9:
                course_map = COURSES_9_GIRLS if gender == "G" else COURSES_9_BOYS
            elif grade == 10:
                course_map = COURSES_10_GIRLS if gender == "G" else COURSES_10_BOYS
            elif grade == 11:
                course_map = COURSES_11
            else:
                course_map = COURSES_12

            for i in range(gender_count):
                first = random.choice(names)
                last = random.choice(LAST_NAMES)

                courses = []
                for dept in ["religion", "english", "math", "science", "history", "language"]:
                    if dept in course_map:
                        options = course_map[dept]
                        if grade in (11, 12):
                            courses.append(_pick_gender_course(options, gender))
                        else:
                            courses.append(random.choice(options))

                if "elective" in course_map:
                    courses.append(random.choice(course_map["elective"]))

                students.append({
                    "id": str(student_id),
                    "last": last,
                    "first": first,
                    "grade": f"{grade}th Grade",
                    "gender": gender,
                    "courses": courses[:7],
                })
                student_id += 1

    return students


TEACHER_DATA = [
    ("Brugger, Mandy", "Science", "Biology, Chemistry, Physics", "Honors Bio, Honors Physics", 6, 1, "Lab requirement", ""),
    ("Lukes, Kathy", "Math", "Algebra 1, Algebra 2, Geometry", "Honors Algebra", 6, 1, "", ""),
    ("Robinson, Andrew", "History", "World History, US History", "AP World Hist, AP US Hist", 6, 1, "", ""),
    ("Johnson, Tara", "Arts/Rotation", "Art Appreciation, MM, AP Art", "AP Art and Design", 5, 2, "Rotation teacher", "Both pools"),
    ("Dlugosielski, Gina", "Music/Rotation", "Music Tech, Band, Choir", "", 5, 2, "Rotation teacher", "Both pools"),
    ("Carly, PT", "Test Prep/Rotation", "ACT Prep", "", 2, 0, "Rotation teacher, Part-time P3+P5B only", "Lunch at P5A"),
    ("Aronoff, Lisa", "Business/Rotation", "Build Wealth, Business Rotational", "", 5, 2, "Rotation teacher", "Both pools"),
    ("Chen, Wei", "Engineering/Rotation", "Engineering", "", 5, 2, "Rotation teacher", "Girls pool"),
    ("Martinez, Rosa", "Science", "Biology, Chemistry", "Honors Chemistry", 6, 1, "", ""),
    ("Thompson, Mark", "Math", "Pre-Calculus, Calculus, Statistics", "Honors Pre-Calc", 6, 1, "", ""),
    ("Williams, Sarah", "English", "English 9, English 10", "Honors English", 6, 1, "", ""),
    ("Davis, Robert", "English", "American Lit, World Lit", "AP American Lit, AP World Lit", 6, 1, "", ""),
    ("Wilson, James", "History", "US History, Economics", "AP US History, AP Economics", 6, 1, "", ""),
    ("Anderson, Karen", "Science", "Physics, Anatomy, Genetics", "AP Physics", 6, 1, "", ""),
    ("Taylor, Michael", "Math", "Algebra 1, Geometry, Math Analysis", "", 6, 1, "", ""),
    ("Brown, Jennifer", "Language", "Spanish 1, Spanish 2, Spanish 3", "", 6, 1, "", ""),
    ("Garcia, Maria", "Language", "French 1, French 2, French 3", "", 6, 1, "", ""),
    ("Moore, David", "Religion", "Sacred Scripture, Sacraments", "", 6, 1, "", ""),
    ("Lee, Christine", "Religion", "Jesus the Redeemer, World Religion", "", 6, 1, "", ""),
    ("Harris, Kevin", "PE", "Physical Education, Health", "", 6, 1, "", ""),
    ("Clark, Amanda", "Science", "Chemistry, AP Chemistry", "AP Chemistry", 6, 1, "Lab requirement", ""),
    ("Lewis, Brian", "Computer Science", "Computer Science, AP Computer Science", "AP Computer Science", 6, 1, "", ""),
    ("Walker, Michelle", "English", "English 9, Honors English 9, English 10, Honors English 10", "Honors English", 6, 1, "", ""),
    ("Hall, Steven", "Math", "Algebra 2, Honors Algebra 2, College Calculus", "Honors Algebra 2", 6, 1, "", ""),
    ("Allen, Patricia", "History", "World History, AP World History", "AP World History", 6, 1, "", ""),
    ("Young, Daniel", "PE", "Physical Education, Health", "", 6, 1, "", ""),
    ("King, Laura", "Arts", "Art Appreciation, Art 2, Studio Art", "", 6, 1, "", ""),
    ("Wright, Timothy", "Music", "Band, Band 2, Choir", "", 6, 1, "", ""),
    ("Lopez, Ana", "Language", "Spanish 2, Spanish 3, Polish", "", 6, 1, "", ""),
    ("Hill, Gregory", "Religion", "Sacraments, Advanced Liturgy", "", 6, 1, "", ""),
    ("Scott, Nicole", "Business", "Business, Intro Business, Speech", "", 6, 1, "", ""),
    ("Adams, Richard", "Psychology", "Psychology, Psychology 2", "", 6, 1, "", ""),
    ("Anwar, Farid", "Computer Science", "Computer Science, Multimedia", "", 6, 1, "", ""),
    ("Nelson, Beth", "English", "American Lit, World Lit, AP World Lit", "AP World Lit", 6, 1, "", ""),
    ("Patel, Sunita", "Science", "Biology, Anatomy", "", 6, 1, "", ""),
    ("Rivera, Carlos", "Math", "Algebra 1, Algebra 2, Geometry", "", 6, 1, "", ""),
    ("Kim, Susan", "English", "English 9, English 10, World Lit", "", 6, 1, "", ""),
    ("O'Brien, Patrick", "History", "World History, US History, Economics", "", 6, 1, "", ""),
    ("Santos, Maria", "Language", "Spanish 1, Spanish 2", "", 6, 1, "", ""),
    ("Mitchell, Karen", "Religion", "Sacred Scripture, Sacraments, Jesus the Redeemer", "", 6, 1, "", ""),
    ("Cooper, James", "Science", "Chemistry, Physics", "", 6, 1, "", ""),
    ("Reed, Stephanie", "Math", "Geometry, Pre-Calculus, Statistics", "", 6, 1, "", ""),
    ("Turner, William", "History", "World History, AP World History, Economics", "AP World History", 6, 1, "", ""),
    ("Phillips, Diana", "English", "American Lit, English 10, English 9", "Honors English", 6, 1, "", ""),
    ("Campbell, Thomas", "Science", "Biology, Chemistry, Genetics", "Honors Biology", 6, 1, "", ""),
    ("Parker, Lisa", "Math", "Algebra 2, College Calculus, Math Analysis", "", 6, 1, "", ""),
    ("Evans, Robert", "Religion", "World Religion, Advanced Liturgy, Sacraments", "", 6, 1, "", ""),
    ("Edwards, Amy", "Language", "French 1, French 2, French 3, Polish", "", 6, 1, "", ""),
    ("Collins, Derek", "PE", "Physical Education, Health", "", 6, 1, "", ""),
    ("Stewart, Rachel", "Arts", "Art Appreciation, Art 2, Computer Science", "", 6, 1, "", ""),
]

ROOM_DATA = [
    ("101", "Standard Classroom", 24, "", "All", ""),
    ("102", "Standard Classroom", 24, "", "All", ""),
    ("103", "Standard Classroom", 24, "", "All", ""),
    ("104", "Standard Classroom", 24, "", "All", ""),
    ("105", "Standard Classroom", 24, "", "All", ""),
    ("106", "Standard Classroom", 24, "", "All", ""),
    ("107", "Standard Classroom", 24, "", "All", ""),
    ("108", "Standard Classroom", 24, "", "All", ""),
    ("109", "Standard Classroom", 24, "", "All", ""),
    ("110", "Standard Classroom", 24, "", "All", ""),
    ("201", "Standard Classroom", 24, "", "All", ""),
    ("202", "Standard Classroom", 24, "", "All", ""),
    ("203", "Standard Classroom", 24, "", "All", ""),
    ("204", "Standard Classroom", 24, "", "All", ""),
    ("205", "Standard Classroom", 24, "", "All", ""),
    ("206", "Standard Classroom", 24, "", "All", ""),
    ("207", "Standard Classroom", 24, "", "All", ""),
    ("208", "Standard Classroom", 24, "", "All", ""),
    ("Lab 1", "Science Lab", 24, "Fume hood, lab benches", "All", "Bio/Chem"),
    ("Lab 2", "Science Lab", 24, "Lab benches", "All", "Physics/Anatomy"),
    ("Lab 3", "Science Lab", 20, "AP Lab setup", "All", "AP Chem/AP Physics"),
    ("Gym A", "Gym", 60, "Full court", "All", "PE classes"),
    ("Gym B", "Gym", 30, "Half court", "All", "PE/Health"),
    ("Art Studio", "Art Room", 20, "Easels, kiln", "All", "Johnson primary"),
    ("Music Hall", "Music Room", 30, "Risers, piano", "All", "Dlugosielski primary"),
    ("Computer Lab", "Computer Lab", 24, "24 workstations", "All", "Anwar/Lewis primary"),
    ("Library", "Library", 30, "Research stations", "All", "Shared"),
    ("Chapel", "Chapel", 300, "Liturgy space", "P4", "SEM only"),
]

CONSTRAINT_DATA = [
    ("Rotation Period", "Girls 9th Grade", "Section A = P1, Section B = P2", "Built-in", "LOCKED"),
    ("Rotation Period", "Boys 10th Grade", "Section A = P3, Section B = P5B", "Built-in", "LOCKED"),
    ("Lunch Time", "All 9-10 students", "P5A", "Built-in", "LOCKED"),
    ("Lunch Time", "All 11-12 students", "P5B", "Built-in", "LOCKED"),
    ("SEM Period", "All", "P4", "Built-in", "LOCKED"),
    ("Teacher Overload", "Brugger, Mandy", "7 teaching periods OK", "Principal + CFO", "APPROVED 2026-06-05"),
    ("Capacity Expansion", "AP World Hist", "24 → 30", "Department Head", "APPROVED 2026-06-05"),
    ("Room Constraint", "Art Studio", "Johnson primary use", "Dean", "LOCKED"),
    ("Room Constraint", "Music Hall", "Dlugosielski primary use", "Dean", "LOCKED"),
]


def generate_course_catalog(students: list[dict]) -> list[dict]:
    course_counts: dict[str, int] = {}
    for s in students:
        for c in s["courses"]:
            course_counts[c] = course_counts.get(c, 0) + 1

    dept_map = {
        "scrip": "Religion", "sacra": "Religion", "jesus": "Religion", "relig": "Religion",
        "litur": "Religion",
        "eng": "English", "lit": "English",
        "alg": "Math", "geo": "Math", "calc": "Math", "pre-calc": "Math",
        "stats": "Math", "math": "Math",
        "bio": "Science", "chem": "Science", "phys": "Science", "anat": "Science",
        "genet": "Science",
        "hist": "History", "econ": "History",
        "spanish": "Language", "french": "Language", "polish": "Language",
        "art": "Arts", "music": "Arts", "band": "Arts",
        "pe": "PE", "health": "PE",
        "comp": "Computer Science", "multi": "Computer Science",
        "psych": "Psychology",
        "business": "Business", "speech": "Business", "build": "Business",
        "intro bus": "Business",
    }

    import math

    catalog = []
    for code, count in sorted(course_counts.items()):
        lower = code.lower()
        dept = "General"
        for key, d in dept_map.items():
            if key in lower:
                dept = d
                break

        grade = 0
        if any(g in lower for g in ["9", "scrip", "alg 1", "bio", "french 1", "spanish 1"]):
            grade = 9
        elif any(g in lower for g in ["10", "sacra", "geo", "chem", "french 2", "spanish 2"]):
            grade = 10
        elif any(g in lower for g in ["11", "jesus", "amer", "pre-calc", "us hist", "french 3", "spanish 3"]):
            grade = 11
        elif any(g in lower for g in ["12", "world relig", "world lit", "calc", "college", "econ", "stats", "litur"]):
            grade = 12

        gender = ""
        if code.startswith("G "):
            gender = "G"
        elif code.endswith(" B") or " B " in code:
            gender = "B"

        flag = ""
        if "AP " in code or "AP " in code.upper():
            flag = "AP"
        elif "Hon " in code or "HON " in code.upper():
            flag = "Honors"

        cap = 24
        if "pe" in lower or "health" in lower:
            cap = 30
        elif "art" in lower:
            cap = 20
        elif "lab" in lower:
            cap = 24

        min_sec = max(1, math.ceil(count / cap))

        catalog.append({
            "code": code,
            "name": code,
            "department": dept,
            "grade": grade,
            "gender": gender,
            "min_sections": min_sec,
            "capacity": cap,
            "flag": flag,
        })

    return catalog


def write_xlsx(students: list[dict], catalog: list[dict], output_path: str):
    wb = openpyxl.Workbook()

    # Sheet 1: Student Selections
    ws1 = wb.active
    ws1.title = "1. Student Selections"
    headers = ["Student ID", "Last Name", "First Name", "Grade", "Gender",
               "Course 1", "Course 2", "Course 3", "Course 4",
               "Course 5", "Course 6", "Course 7"]
    for col, h in enumerate(headers, 1):
        ws1.cell(row=1, column=col, value=h)

    for row, s in enumerate(students, 2):
        ws1.cell(row=row, column=1, value=s["id"])
        ws1.cell(row=row, column=2, value=s["last"])
        ws1.cell(row=row, column=3, value=s["first"])
        ws1.cell(row=row, column=4, value=s["grade"])
        ws1.cell(row=row, column=5, value=s["gender"])
        for i, c in enumerate(s["courses"][:7]):
            ws1.cell(row=row, column=6 + i, value=c)

    # Sheet 2: Course Catalog
    ws2 = wb.create_sheet("2. Course Catalog")
    headers = ["Course Code", "Course Name", "Department", "Grade Level",
               "Gender Restriction", "Min Sections", "Standard Capacity", "Honors/AP Flag"]
    for col, h in enumerate(headers, 1):
        ws2.cell(row=1, column=col, value=h)

    for row, c in enumerate(catalog, 2):
        ws2.cell(row=row, column=1, value=c["code"])
        ws2.cell(row=row, column=2, value=c["name"])
        ws2.cell(row=row, column=3, value=c["department"])
        ws2.cell(row=row, column=4, value=c["grade"])
        ws2.cell(row=row, column=5, value=c["gender"])
        ws2.cell(row=row, column=6, value=c["min_sections"])
        ws2.cell(row=row, column=7, value=c["capacity"])
        ws2.cell(row=row, column=8, value=c["flag"])

    # Sheet 3: Teacher Roster
    ws3 = wb.create_sheet("3. Teacher Roster")
    headers = ["Teacher Name", "Department", "Subject Specialties",
               "Certifications (AP/Honors/etc)", "Max Teaching Periods",
               "Required PREP Periods", "Special Constraints", "Notes"]
    for col, h in enumerate(headers, 1):
        ws3.cell(row=1, column=col, value=h)

    for row, t in enumerate(TEACHER_DATA, 2):
        for col, val in enumerate(t, 1):
            ws3.cell(row=row, column=col, value=val)

    # Sheet 4: Room Inventory
    ws4 = wb.create_sheet("4. Room Inventory")
    headers = ["Room Number", "Room Type", "Capacity", "Special Features",
               "Available Periods", "Notes"]
    for col, h in enumerate(headers, 1):
        ws4.cell(row=1, column=col, value=h)

    for row, r in enumerate(ROOM_DATA, 2):
        for col, val in enumerate(r, 1):
            ws4.cell(row=row, column=col, value=val)

    # Sheet 5: Constraints
    ws5 = wb.create_sheet("5. Constraints")
    headers = ["Constraint Type", "Subject/Teacher", "Detail",
               "Approval Required From", "Status"]
    for col, h in enumerate(headers, 1):
        ws5.cell(row=1, column=col, value=h)

    for row, c in enumerate(CONSTRAINT_DATA, 2):
        for col, val in enumerate(c, 1):
            ws5.cell(row=row, column=col, value=val)

    wb.save(output_path)
    print(f"Generated sample data: {output_path}")
    print(f"  Students: {len(students)}")
    print(f"  Courses: {len(catalog)}")
    print(f"  Teachers: {len(TEACHER_DATA)}")
    print(f"  Rooms: {len(ROOM_DATA)}")


def main():
    import sys
    output = sys.argv[1] if len(sys.argv) > 1 else "olsm_sample_data.xlsx"
    random.seed(42)
    students = generate_students(847)
    catalog = generate_course_catalog(students)
    write_xlsx(students, catalog, output)


if __name__ == "__main__":
    main()
